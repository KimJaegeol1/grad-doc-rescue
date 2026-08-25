# -*- coding: utf-8 -*-
"""
smoke_general — 서울대 것이 아닌 엑셀에서도 도는가 (범용성)

여태 고친 것이 전부 **서울대 한 학교, 한 시트**를 돌려 보다 나왔다.  그러면
"서울대에만 맞춘 것 아니냐" 는 물음이 남는다.  그래서 서울대와 **일부러
전부 다르게** 생긴 엑셀을 손으로 지어 넣고 끝까지 돌린다.

    서울대                          여기 (지어낸 학교)
    ────────────────────────────    ──────────────────────────────
    시트 이름  「⑤ 인턴십」           「인턴십 실적」   번호가 아예 없다
    머리글     2행                   4행           위에 안내문 두 줄
    참여학생명  O2:AE2 병합 · 칸마다 하나  한 칸에 쉼표로  정반대 꼴
    기업명     「기업명」              「참여기관」     낱말장의 별칭 → 찾는다
    인원       「참여인원」            「인원」        낱말장의 별칭 → 찾는다
    시간       「시간(00H)」           「총 소요시수」  **아무 데도 없는 말** → 알린다
    별첨1      명단 61명 있다           **아예 없다**
    줄 수      22                    5             배치가 한 번에 끝난다

이 파일에는 실제 학생 이름이 없다.  전부 지어낸 것이라 함께 묶어도 된다.
**남의 컴퓨터에서도 이 시험만은 언제나 돈다.**

    python tools/smoke_general.py
"""
from __future__ import annotations

import json
import shutil
import sys
import zipfile
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE.parent))

from openpyxl import Workbook, load_workbook                       # noqa: E402
from openpyxl.styles import PatternFill                            # noqa: E402

from core import ask as A, book as B, flow as F, spec as S         # noqa: E402

일터 = HERE.parent / "workspace" / "_general"
시트 = "⑤인턴십"
검증칠 = PatternFill("solid", fgColor="FCE4D6")
탈 = []


def 봄(ok, 글, 덧=""):
    print(f"  {'✓' if ok else '❌'} {글}" + (f"   {덧}" if 덧 else ""))
    if not ok:
        탈.append(글)


def 칸(제목):
    print("\n" + "═" * 74)
    print(제목)
    print("═" * 74)


# ══════════════════════════════════════════════════════════════
def 지어낸엑셀(경로: Path, 검증열들: list):
    """서울대와 닮은 데가 없는 성과관리현황을 짓는다."""
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("표지")                       # 엉뚱한 시트도 하나 둔다
    ws = wb.create_sheet("인턴십 실적")            # ★ 번호 기호가 없다
    ws["A1"] = "○○대학교 녹색융합기술인재 양성사업"   # ★ 머리글 위 안내문 두 줄
    ws["A2"] = "※ 인턴십 실적을 아래에 적어 주십시오"

    #  「참여기관」·「인원」·「참여학생」 은 낱말장에 있는 별칭 → 찾아야 한다
    #  「총 소요시수」 는 **아무 데도 없는 말** → 못 찾고 **알려야** 한다
    자료 = ["순번", "참여기관", "일시", "총 소요시수", "인원", "참여학생"]
    for i, h in enumerate(자료, 1):
        ws.cell(4, i, h)                          # ★ 머리글이 4행
    for j, h in enumerate(검증열들, len(자료) + 1):
        c = ws.cell(4, j, h)
        c.fill = 검증칠                            # 검사 칸은 색으로 알아본다
    ws.cell(4, len(자료) + len(검증열들) + 1, "비고").fill = 검증칠
    ws.cell(4, len(자료) + len(검증열들) + 2, "학교 의견").fill = 검증칠

    줄 = [("1", "한빛환경기술", "25-03-04 ~ 25-03-14", "40", "3",
           "가나다, 라마바, 사아자"),                # ★ 한 칸에 쉼표로
          ("2", "푸른하늘연구소", "25-04-07 ~ 25-04-18", "40", "2",
           "차카타 · 파하가"),                       # ★ 가운뎃점도 섞는다
          ("3", "맑은물㈜", "25-05-12 ~ 25-05-23", "40", "4",
           "나다라, 마바사, 아자차, Kwame Osei Mensah"),   # ★ 긴 외국 이름
          ("4", "대기측정기술원", "25-06-02 ~ 25-06-13", "40", "1", "타파하"),
          ("5", "", "", "", "", "")]                # ★ 순번만 적힌 줄 — 우리가
                                                    #   아는 열엔 값이 없다
    for r, 값 in enumerate(줄, 5):
        for c, v in enumerate(값, 1):
            ws.cell(r, c, v)
    wb.save(경로)
    wb.close()
    return len(줄)


def docx(path: Path, 줄들):
    body = "".join(f'<w:p><w:r><w:t xml:space="preserve">{x}</w:t></w:r></w:p>'
                   for x in 줄들)
    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "w") as z:
        z.writestr("[Content_Types].xml",
                   '<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats'
                   '.org/package/2006/content-types"><Default Extension="xml" '
                   'ContentType="application/xml"/><Override PartName="/word/'
                   'document.xml" ContentType="application/vnd.openxmlformats-'
                   'officedocument.wordprocessingml.document.main+xml"/></Types>')
        z.writestr("_rels/.rels",
                   '<?xml version="1.0"?><Relationships xmlns="http://schemas.'
                   'openxmlformats.org/package/2006/relationships"><Relationship '
                   'Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument'
                   '/2006/relationships/officeDocument" Target="word/document.xml"/>'
                   '</Relationships>')
        z.writestr("word/document.xml",
                   '<?xml version="1.0"?><w:document xmlns:w="http://schemas.'
                   'openxmlformats.org/wordprocessingml/2006/main"><w:body>'
                   + body + "</w:body></w:document>")


# ══════════════════════════════════════════════════════════════
# 모형 흉내 — 이름이 아니라 **표**로 답한다 (규격이 그것만 받는다)
# ══════════════════════════════════════════════════════════════
class _응답:
    def __init__(self, d):
        self.usage = type("u", (), {"prompt_tokens": 1234, "completion_tokens": 567})()
        self.choices = [type("c", (), {"message": type(
            "m", (), {"content": json.dumps(d, ensure_ascii=False)})()})()]


def _답(글: str, schema: dict) -> dict:
    줄꼴 = schema["properties"]["줄"]
    물은줄 = [int(k[1:]) for k in 줄꼴["properties"]]
    줄칸 = 줄꼴["properties"][f"r{물은줄[0]}"]["properties"]
    값이름 = list(줄칸["값"]["properties"])
    판정키 = [k for k in 줄칸 if k.startswith("v") and k[1:].isdigit()]

    파일 = []
    for ln in 글.splitlines():
        if ln.startswith("【file_"):
            표, _, 이름 = ln[1:].partition("】")
            파일.append((표.strip(), 이름.strip()))

    맺 = {5: ("한빛", "한빛환경기술", "2025-03-04", "2025-03-14",
              ["가나다", "라마바", "사아자"]),
          6: ("푸른", "푸른하늘연구소", "2025-04-07", "2025-04-18",
              ["차카타", "파하가"])}

    줄 = {}
    for n in 물은줄:
        낱말, 기업, s, e, 이름 = 맺.get(n, ("", "", "", "", []))
        f = [t for t, nm in 파일 if 낱말 and 낱말 in nm]
        값 = {k: "" for k in 값이름}
        for k, v in (("기업명", 기업), ("인턴기간_시작일", s), ("인턴기간_종료일", e),
                     (S.인원칸, str(len(이름)) if 이름 else ""), ("총시간", "40" if f else "")):
            if k in 값:
                값[k] = v
        if S.명단칸 in 값:
            값[S.명단칸] = {"상태": "찾음" if 이름 else "명단없음", "이름": 이름}
        if "문서종류" in 값:
            값["문서종류"] = ["인턴십 계획서"] if f else []
        한 = {"행": n, "맺은파일": f, "확신": "높음" if f else "못맺음",
              "맺은근거": "참여기관·기간 일치" if f else "폴더에 그 기관 서류가 없음",
              "값": 값, "메모": ""}
        for i, k in enumerate(판정키):
            한[k] = {"검증열": k, "판정": ("O" if f and i < 2 else "확인 불가"),
                     "출처": (f[0] if f and i < 2 else "없음"),
                     "근거": ("p.1 표제에서 봄" if f and i < 2 else "맺은 서류가 없음"),
                     "비고": ""}
        줄[f"r{n}"] = 한
    쓴 = {y for x in 줄.values() for y in x["맺은파일"]}
    return {"줄": 줄, "안맺힌파일": [{"파일": t, "왜": "다른 시트 증빙인 듯하다"}
                                    for t, _n in 파일 if t not in 쓴]}


class _가짜:
    class chat:
        class completions:
            @staticmethod
            def create(**kw):
                return _응답(_답(kw["messages"][1]["content"],
                                kw["response_format"]["json_schema"]["schema"]))


# ══════════════════════════════════════════════════════════════
def main() -> int:
    import os
    if 일터.exists():
        shutil.rmtree(일터)
    일터.mkdir(parents=True)

    sp = S.load(F.기준찾기())
    검증열들 = [c["검증열"] for c in sp.항목(시트)]
    책 = 일터 / "○○대_성과관리현황.xlsx"
    줄수 = 지어낸엑셀(책, 검증열들)

    증빙 = 일터 / "증빙"
    docx(증빙 / "인턴십_한빛환경기술_계획서.docx", [
        "인턴십 계획서", "참여기관: 한빛환경기술",
        "기간: 2025년 03월 04일 ~ 2025년 03월 14일",
        "참여학생: 가나다, 라마바, 사아자", "총 40시간"])
    docx(증빙 / "인턴십_푸른하늘연구소.docx", [
        "인턴십 계획서", "참여기관: 푸른하늘연구소",
        "기간: 2025년 04월 07일 ~ 2025년 04월 18일",
        "참여학생: 차카타, 파하가", "총 40시간"])
    docx(증빙 / "엉뚱한_학술발표_초록.docx", ["초록", "○○ 저감 연구", "춘계학술대회"])

    A._client = lambda: _가짜()
    os.environ.setdefault("OPENAI_API_KEY", "시험용")

    칸("① 서울대와 닮은 데가 없는 엑셀을 끝까지 돌린다")
    print("  시트 이름에 번호가 없고 · 머리글이 4행이고 · 참여학생이 한 칸에")
    print("  쉼표로 들어 있고 · 열 이름이 죄다 별칭 쪽이고 · 별첨1 이 아예 없다.\n")
    res = F.검사(책, 시트, 증빙, OCR=False, workers=2, log=lambda m, lv="info": None)
    봄(Path(res["검사결과"]).exists(), "검사 결과가 나왔다", Path(res["검사결과"]).name)

    칸("② 번호 없는 시트 · 4행 머리글 · 별칭 열을 찾아냈는가")
    bk = B.Book(res["검사결과"], 시트)
    지도, 못, 셈 = B.열지도(bk, sp.항목(시트))
    봄(bk.ws.title == "인턴십 실적", "번호가 없어도 시트를 찾는다", bk.ws.title)
    봄(bk.머리글줄 == 4, "머리글이 2행이 아니어도 찾는다", f"{bk.머리글줄}행")
    봄(set(지도) >= {"참여학생명", "일시", "기업명", "참여학생 수"},
      "★ 별칭으로 적힌 열을 낱말장으로 찾아냈다 (참여기관·인원·참여학생)",
      " · ".join(f"{k}={len(v)}칸" for k, v in 지도.items()))
    봄(all(len(v) == 1 for v in 지도.values()),
      "★ 병합이 없으니 한 칸씩만 잡는다 — 서울대 것에 맞춘 게 아니다")
    줄들 = bk.줄들(지도)
    봄(bk.별첨1() == [], "별첨1 이 없어도 멈추지 않는다")

    print("\n  ── 그래도 못 찾는 말이 있다.  **조용히 넘어가지 않는다** ──")
    print("  이 학교는 시간을 「총 소요시수」 라 적었다.  낱말장에 없는 말이다.")
    print("  코드가 할 일은 그걸 **알리는 것**이고, 낱말을 늘리는 것은 정본이 할 일이다.\n")
    봄(any("시간" in x["검증열"] for x in 못),
      "못 찾은 열을 집어서 알린다", " · ".join(x["검증열"] for x in 못) or "(없음)")
    봄(any("건너뛰었습니다" in x for x in bk.탈),
      "★ 그 바람에 사라진 줄도 알린다 — 조용히 없어지는 것이 제일 나쁘다",
      next((x[:70] for x in bk.탈 if "건너뛰" in x), ""))
    봄(len(줄들) == 줄수 - 1,
      f"값이 있는 {줄수 - 1}줄을 읽었다", f"{len(줄들)}줄")
    print("\n  ── 열도 조용히 넘기지 않는다 (개요 제0조 ②) ──")
    print("  「순번」·「총 소요시수」 처럼 어느 항목도 안 가져간 열이 있으면 알린다.")
    print("  낱말장을 상상으로 늘리는 대신, 이 알림을 보고 **그 한 줄만** 더한다.\n")
    임자없는 = 셈.get("임자없는열") or []
    봄("총 소요시수" in 임자없는,
      "★ 뜻을 모르는 머리글을 집어서 알린다 — 알아본 척하지 않는다",
      " · ".join(임자없는) or "(없음)")
    봄("참여기관" not in 임자없는 and "인원" not in 임자없는,
      "알아본 열은 알림에 안 뜬다")
    print()
    print("  ▌ 원칙:  비어 있지 않은 줄은 반드시 셋 중 하나로 끝난다 —")
    print("           검수됨 · 검수 대상 아님 · 구조를 이해하지 못해 보류.")
    print("           열도 같다.  조용히 사라지는 것은 없다.")
    bk.close()

    칸("③ 한 칸에 쉼표로 든 이름도 제대로 센다")
    from core import calc as C
    항 = next(c for c in sp.항목(시트) if c["검증열"] == "참여인원 일치")
    셈 = {}
    for r in 줄들:
        v, _b, 근거 = C.재기(항, r["값"], {}, 맺은파일=["01.docx"])
        셈[v or "공란"] = 셈.get(v or "공란", 0) + 1
        print(f"     r{r['행']}  {v or '공란':<6} {근거}")
    봄(셈.get("X", 0) == 0, "★ 거짓 X 가 없다")
    봄(셈.get("O", 0) == 4, "값이 있는 네 줄이 다 O 다 (긴 외국 이름 포함)")

    칸("④ 검사 칸과 비고가 채워졌는가")
    wb = load_workbook(res["검사결과"])
    ws = wb["인턴십 실적"]
    쓴칸 = sum(1 for r in range(5, 5 + 줄수) for c in range(7, ws.max_column + 1)
               if ws.cell(r, c).value not in (None, ""))
    봄(쓴칸 > 0, f"검사 칸에 {쓴칸}개가 쓰였다")
    봄(str(ws.cell(5, 7).value or "") == "O", "맺은 줄의 첫 검사 칸이 O 다",
      str(ws.cell(5, 7).value))
    봄(str(ws.cell(5, 1).value) == "1" and str(ws.cell(5, 2).value) == "한빛환경기술",
      "★ 자료 열은 한 글자도 안 건드렸다")
    wb.close()

    칸("⑤ 정본이 코드가 아는 이름을 쓰고 있는가")
    print("  「참여명단」·「참여인원수」 는 코드가 뜻을 아는 딱 둘뿐인 이름이다.")
    print("  정본에서 이 이름을 바꾸면 명단 대조가 조용히 죽는다.\n")
    봄(S.명단칸 in sp.뽑을값들(시트) and S.인원칸 in sp.뽑을값들(시트),
      "정본과 코드의 이름이 맞는다", f"{S.명단칸} · {S.인원칸}")
    봄(not [x for x in sp.탈 if "코드는" in x], "약속이 어긋났다는 알림이 없다")

    print()
    print("═" * 74)
    if 탈:
        print(f"❌ 어긋난 곳 {len(탈)}건")
        for x in 탈:
            print(f"   · {x}")
    else:
        print("✅ 서울대 것이 아닌 엑셀에서도 그대로 돈다 — 데이터에 맞춘 코드가 아니다")
    print("═" * 74)
    return 1 if 탈 else 0


if __name__ == "__main__":
    raise SystemExit(main())
