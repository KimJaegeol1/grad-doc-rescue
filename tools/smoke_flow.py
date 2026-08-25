# -*- coding: utf-8 -*-
"""
smoke_flow — 폴더를 넣으면 검사 칸이 채워지는가 (모형 자리에 가짜를 앉히고)

진짜 OpenAI 를 부르지 않는다.  **모형 자리에만 가짜를 앉히고 나머지는 전부
진짜 코드**로 돈다 — 검수기준 읽기 · 학교 엑셀 읽기 · 파일 열기 · 프롬프트
조립 · 답 검사 · 코드 셈 · 검사 칸 쓰기 · 추출 내역.

학교 엑셀은 **서울대가 실제로 낸 것**을 쓴다 (머리글에 줄바꿈이 있고,
⑧학술발표가 「⑦ 학술발표」 로 잘못 매겨져 있는 그 파일).

    python tools/smoke_flow.py
"""
from __future__ import annotations

import json
import shutil
import sys
import zipfile
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE.parent))

from openpyxl import load_workbook                                # noqa: E402

from core import ask as A, book as B, flow as F, prompt as P, spec as S  # noqa: E402

# 시험대 엑셀.  학생 이름이 들어 있어 함께 묶지 않는다 —
# 없으면 조용히 건너뛴다.  다른 성과관리현황을 넘겨 써도 된다.
#     python tools/smoke_flow.py <성과관리현황.xlsx>
원본 = Path(sys.argv[1] if len(sys.argv) > 1 else
            "/root/.claude/uploads/e2f6df8d-7996-5b70-ad90-088559601eb3/"
            "4aa41b31-_______________.xlsx")
일터 = HERE.parent / "workspace" / "_flow"
시트 = "⑤인턴십"
탈 = []


def 봄(ok, 글, 덧=""):
    print(f"  {'✓' if ok else '❌'} {글}" + (f"   {덧}" if 덧 else ""))
    if not ok:
        탈.append(글)


def 칸(제목):
    print("\n" + "═" * 74)
    print(제목)
    print("═" * 74)


# ══════════════════════════════════════════════════════════════
# 시험용 증빙 — 글만 있으면 되므로 가장 단순한 docx
# ══════════════════════════════════════════════════════════════
def docx(path: Path, 줄들):
    body = "".join(f'<w:p><w:r><w:t xml:space="preserve">{x}</w:t></w:r></w:p>'
                   for x in 줄들)
    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "w") as z:
        z.writestr("[Content_Types].xml",
                   '<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats'
                   '.org/package/2006/content-types"><Default Extension="xml" '
                   'ContentType="application/xml"/><Override PartName="/word/'
                   'document.xml" ContentType="application/vnd.openxmlformats-'
                   'officedocument.wordprocessingml.document.main+xml"/></Types>')
        z.writestr("_rels/.rels",
                   '<?xml version="1.0"?><Relationships xmlns="http://schemas.'
                   'openxmlformats.org/package/2006/relationships"><Relationship '
                   'Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument'
                   '/2006/relationships/officeDocument" Target="word/document.xml"/>'
                   '</Relationships>')
        z.writestr("word/document.xml",
                   '<?xml version="1.0"?><w:document xmlns:w="http://schemas.'
                   'openxmlformats.org/wordprocessingml/2006/main"><w:body>'
                   + body + "</w:body></w:document>")


def 준비():
    if 일터.exists():
        shutil.rmtree(일터)
    일터.mkdir(parents=True)
    책 = 일터 / "서울대검증문서.xlsx"
    shutil.copy2(원본, 책)

    증빙 = 일터 / "제출자료" / "2024" / "인턴십"       # 폴더 안 폴더
    docx(증빙 / "01_인턴십_과학원.docx", [
        "인턴십 계획서", "기업명: 국립환경과학원",
        "인턴기간: 24년 06월 03일 ~ 24년 06월 16일", "참여학생: 서혜빈", "총 40시간",
        "───── 다음 쪽 ─────", "결과보고서 요약서", "대기오염 측정망 자료 분석 보조",
        "───── 다음 쪽 ─────", "세부 결과보고서", "1일차 오리엔테이션 …", "현장 사진 3매"])
    docx(증빙 / "02_인턴십_그리디그리드.docx", [
        "인턴십 계획서", "기업명: 그리디그리드",
        "인턴기간: 24년 06월 03일 ~ 24년 06월 19일", "참여학생: 김은진", "총 40시간"])
    docx(일터 / "제출자료" / "09_딴시트_학술발표.docx", [
        "초록", "미세먼지 저감을 위한 …", "한국환경분석학회 2024 춘계학술대회"])
    return 책, 일터 / "제출자료"


# ══════════════════════════════════════════════════════════════
# 모형 흉내 — 스키마를 보고 사람이 낼 법한 답을 낸다
# ══════════════════════════════════════════════════════════════
class _응답:
    def __init__(self, d):
        self.usage = type("u", (), {"prompt_tokens": 41234, "completion_tokens": 5678})()
        self.choices = [type("c", (), {"message": type(
            "m", (), {"content": json.dumps(d, ensure_ascii=False)})()})()]


def _답(글: str, schema: dict) -> dict:
    # ★ 줄은 이제 r3·r4 칸으로 온다.  물은 줄이 곧 규격의 required 다 —
    #   가짜도 거기 적힌 줄만, 그리고 **전부** 답해야 한다.
    줄꼴 = schema["properties"]["줄"]
    물은줄 = [int(k[1:]) for k in 줄꼴.get("properties", {})]
    줄칸 = 줄꼴["properties"][f"r{물은줄[0]}"]["properties"]
    값이름 = list(줄칸["값"]["properties"])
    판정키 = [k for k in 줄칸 if k.startswith("v") and k[1:].isdigit()]

    # 프롬프트에서 파일 표를 읽어 온다 (사람이 하듯)
    파일 = []
    # ★ 파일은 이제 【file_01】 이름 꼴로 실린다.  답에 쓸 수 있는 것은 **표**뿐이다
    #   (규격이 enum 으로 막는다).  가짜도 진짜와 똑같이 표로 답해야 한다.
    for ln in 글.splitlines():
        if ln.startswith("【file_"):
            표, _, 이름 = ln[1:].partition("】")
            파일.append((표.strip(), 이름.strip()))

    def 값(기업, 시작, 끝, 이름):
        v = {f: "" for f in 값이름}
        if "참여명단" in v:
            v["참여명단"] = {"상태": "찾음" if 이름 else "명단없음",
                             "이름": [이름] if 이름 else []}
        if "문서종류" in v:
            v["문서종류"] = ["인턴십 계획서", "결과보고서 요약서"]
        for k, x in (("기업명", 기업), ("인턴기간_시작일", 시작),
                     ("인턴기간_종료일", 끝), ("총시간", "40"),
                     ("참여인원수", f"참여학생 {1 if 이름 else 0}명")):
            if k in v:
                v[k] = x
        return v

    # ★ 파일을 **차례로** 집으면 안 된다.  캐시 이름이 내용 해시라 프롬프트에
    #   실리는 차례가 이름 차례와 다르다.  사람(과 진짜 모형)이 하듯 **내용으로** 고른다.
    def 골라(낱말):
        return [t for t, 이름 in 파일 if 낱말 in 이름]

    맺 = {3: (골라("과학원"), "국립환경과학원", "24년 06월 03일",
              "24년 06월 16일", "서혜빈", "높음"),
          4: (골라("그리디"), "그리디그리드", "24년 06월 03일",
              "24년 06월 19일", "김은진", "높음")}

    줄 = {}
    for n in 물은줄:
        f, 기업, s, e, 이름, 확신 = 맺.get(n, ([], "", "", "", "", "못맺음"))
        한 = {"행": n, "맺은파일": f,
              "맺은근거": "기업명·기간·학생 일치" if f else "폴더에서 못 찾음",
              "확신": 확신, "값": 값(기업, s, e, 이름), "메모": ""}
        for i, k in enumerate(판정키):
            한[k] = {"검증열": k,
                     "판정": ("O" if f and i < 2 else "확인 불가"),
                     "출처": (f[0] if f and i < 2 else "없음"),
                     "근거": ("p.1 표제에서 봄" if f and i < 2 else "맺은 서류가 없음"),
                     "비고": ""}
        줄[f"r{n}"] = 한
    쓴것 = {y for fs, *_r in 맺.values() for y in fs}
    남 = [t for t, _이름 in 파일 if t not in 쓴것]
    return {"줄": 줄,
            "안맺힌파일": [{"파일": x, "왜": "이 시트 실적이 아닌 듯하다"} for x in 남]}


def _빈답만들기(schema: dict) -> dict:
    """모형이 어깨만 으쓱한 답 — 규격은 다 맞지만 알맹이가 없다.

    실제로 서울대 2번째 배치가 이렇게 왔다.  줄 칸은 다 채우고, 안은 다 비웠다.
    """
    줄꼴 = schema["properties"]["줄"]
    줄 = {}
    for k, v in 줄꼴.get("properties", {}).items():
        칸 = v["properties"]
        한 = {"행": int(k[1:]), "맺은파일": [], "맺은근거": "",
              "확신": "못맺음", "메모": "",
              "값": {f: ({"상태": "명단없음", "이름": []} if f == "참여명단"
                         else ([] if 칸["값"]["properties"][f]["type"] == "array" else ""))
                     for f in 칸["값"]["properties"]}}
        for vk in [x for x in 칸 if x.startswith("v") and x[1:].isdigit()]:
            한[vk] = {"검증열": vk, "판정": "확인 불가", "출처": "없음",
                      "근거": "", "비고": ""}
        줄[k] = 한
    return {"줄": 줄, "안맺힌파일": []}


class _가짜:
    """진짜 모형 자리에 앉는다.  `으쓱` 에 담긴 횟수만큼 빈 답을 낸다."""

    으쓱 = 0
    부른수 = 0

    class chat:
        class completions:
            @staticmethod
            def create(**kw):
                글 = kw["messages"][1]["content"]
                sch = kw["response_format"]["json_schema"]["schema"]
                _가짜.부른수 += 1
                if _가짜.으쓱 > 0:
                    _가짜.으쓱 -= 1
                    return _응답(_빈답만들기(sch))
                return _응답(_답(글, sch))


# ══════════════════════════════════════════════════════════════
def main() -> int:
    import os
    if not 원본.exists():
        print(f"시험대 성과관리현황 엑셀이 없습니다 — {원본}")
        print("검사 칸(주황)이 든 엑셀을 인자로 주시면 그것으로 돕니다.  건너뜁니다.")
        return 0
    책, 증빙 = 준비()
    A._client = lambda: _가짜()
    os.environ.setdefault("OPENAI_API_KEY", "시험용")

    칸("① 폴더를 넣고 끝까지 돌린다")
    기록 = []
    res = F.검사(책, 시트, 증빙, workers=2, OCR=False,
                 log=lambda m, lv="info": (기록.append((str(m), lv)),
                                           print("     " + str(m)[:100])
                                           if lv in ("head", "done", "err", "warn")
                                           else None))
    봄(Path(res["검사결과"]).exists(), "검사 결과가 나왔다",
      Path(res["검사결과"]).name)
    봄(Path(res["추출내역"]).exists(), "추출 내역이 나왔다",
      Path(res["추출내역"]).name)
    봄(Path(res["백업"]).exists(), "원본 백업을 떴다", Path(res["백업"]).name)
    # ★ 검사 결과에도 시트 이름이 붙는다.  engines.write 는 쓸 때마다 원본을
    #   다시 복사하므로, 이름이 하나뿐이면 다음 시트가 앞 시트 결과를 지운다.
    봄(Path(res["검사결과"]).stem.endswith(시트),
      "★ 검사 결과 이름 끝에 시트가 붙는다 — 다음 시트가 덮어쓰지 못한다",
      Path(res["검사결과"]).name)
    봄(Path(res["추출내역"]).stem.endswith(시트),
      "추출 내역과 이름 규칙이 같다", Path(res["추출내역"]).name)
    봄(Path(res["검사결과"]).name != Path(res["추출내역"]).name, "둘은 다른 파일이다")
    print(f"     집계 {res['집계']}")

    칸("② 검사 칸이 실제로 채워졌는가")
    wb = load_workbook(res["검사결과"])
    bk = B.Book(res["검사결과"], 시트)
    hr, 검 = bk.머리글줄, bk.검증열
    ws = bk.ws
    쓴칸 = 0
    for r in (3, 4, 5):
        vs = {h: str(ws[f"{c}{r}"].value or "") for h, c in 검.items()}
        쓴칸 += sum(1 for v in vs.values() if v)
        print(f"     {r}줄  " + "  ".join(
            f"{h[:8]}={v or '·'}" for h, v in list(vs.items())[:5]))
    봄(쓴칸 > 0, f"검사 칸에 {쓴칸}개가 쓰였다")

    sp = S.load(F.기준찾기())
    검증열 = [c["검증열"] for c in sp.항목(시트)]
    _ = 0
    계 = next(h for h in 검 if "계획서" in h)
    봄(str(ws[f"{검[계]}3"].value or "") == "O", "3줄 계획서 제출여부 = O")
    시간칸 = next((h for h in 검 if h.strip() == "시간 일치"), None)
    if 시간칸:
        봄(str(ws[f"{검[시간칸]}3"].value or "") == "O",
          "3줄 시간 일치 = O  (코드가 셈 · 엑셀 40 = 증빙 40)")
    비고칸 = next((c for h, c in 검.items() if "비고" in h and "학교" not in h), None)
    if 비고칸:
        글 = str(ws[f"{비고칸}3"].value or "")
        print("     3줄 비고:")
        for ln in (글 or "(빈칸)").splitlines():
            print(f"        {ln[:100]}")
        봄(bool(글.strip()), "비고에 사유가 적혔다")
        봄("[계획서 제출여부]" not in 글,
          "★ O 인 항목은 비고에 안 적는다 — 잘된 것을 적을수록 잘못된 것이 묻힌다")
        봄(all(ln.lstrip().startswith("[") for ln in 글.splitlines() if ln.strip()),
          "토막마다 줄이 선다 — 사유 안의 ' / ' 와 안 헷갈린다")

    # ── ★ 비고는 손볼 것만 적는다 ─────────────────────────
    print("\n  예전엔 O 까지 다 적어서 한 칸이 이랬다 (실제 서울대 3줄) —")
    print("     [계획서 제출여부] [계획서 제출여부] 계획서 1건 / [결과보고서 요약서")
    print("     제출여부] [결과보고서 요약서 제출여부] 요약서 있음 / … / [참여인원")
    print("     일치] 엑셀 5명 / 이름 1명")
    print("  아홉 토막 중 봐야 할 것은 맨 뒤 둘뿐이었다.\n")
    for 판정, 적나, 왜 in (
            ("X", True, "학교에 반송이 나간다 — 왜인지 적어야 한다"),
            ("확인 불가", True, "사람이 봐야 한다 — 왜 못 가렸는지 적어야 한다"),
            ("O", False, "할 일이 없다"),
            ("", False, "판정할 대상이 아니다")):
        t = F._비고("기업명 일치", {"판정": 판정, "비고": "엑셀 A  ↔  증빙 B"})
        봄(bool(t) is 적나,
          f"{판정 or '공란':<6} → {'적는다' if 적나 else '안 적는다'}   {왜}")
    봄(F._비고("계획서 제출여부",
             {"판정": "X", "비고": "[계획서 제출여부] 계획서 1건"}) == "계획서 1건",
      "모형이 되뇐 항목 이름을 뗀다 — engines.write 가 또 붙여 두 번 찍혔다")
    봄(F._비고("일시 일치", {"판정": "X", "비고": "[다른 것] 무엇"}) == "[다른 것] 무엇",
      "제 이름이 아닌 대괄호는 그대로 둔다")

    bk.close()

    칸("③ 자료 열은 안 건드렸는가")
    원 = load_workbook(원본, data_only=True)
    원ws = 원[ws.title]
    같음 = True
    for r in range(3, 12):
        for c in range(1, 20):
            a = 원ws.cell(r, c).value
            b = ws.cell(r, c).value
            if str(a or "") != str(b or "") and B.gl(c) not in set(검.values()):
                같음 = False
                print(f"     ❌ {B.gl(c)}{r}  「{a}」 → 「{b}」")
    봄(같음, "★ 자료 열·학교 의견 열은 한 글자도 안 바뀌었다")
    원.close()

    칸("④ 추출 내역")
    wb2 = load_workbook(res["추출내역"])
    봄(wb2.sheetnames == ["개요", 시트, "판정", "남은 것"], "장 넷",
      " · ".join(wb2.sheetnames))
    개 = "\n".join(str(wb2["개요"].cell(r, 1).value or "")
                   for r in range(1, wb2["개요"].max_row + 1))
    봄("봐야 할 것" in 개, "「봐야 할 것」 이 있다")
    남 = "\n".join(str(wb2["남은 것"].cell(r, c).value or "")
                   for r in range(1, wb2["남은 것"].max_row + 1) for c in (1, 2))
    봄("09_딴시트_학술발표.docx" in 남, "다른 시트 증빙을 알린다")
    for s in res["신호"][:6]:
        print(f"     [{s['수준']}] {s['글'][:88]}")

    칸("⑤ 폴더 안 폴더까지 읽었는가")
    캐시 = 증빙 / "text_cache"
    본 = list(캐시.glob("*.txt"))
    봄(len(본) == 3, f"파일 3개를 다 읽었다 ({len(본)}개) — 2단계 깊이까지")

    칸("⑥ ★ 모형이 어깨만 으쓱하면 다시 묻는다")
    print("  서울대 2번째 배치가 여덟 줄을 통째로 빈 답으로 냈다.  재시도도 아니었고,")
    print("  그 여덟 줄의 서류는 폴더에 다 있었다.  빈 답을 받아 쓰면 그 줄은")
    print("  전부 확인 불가가 되고, 사람은 왜인지도 모른다.\n")
    _가짜.으쓱, _가짜.부른수 = 1, 0
    기록2 = []
    res2 = F.검사(책, 시트, 증빙, workers=2, OCR=False,
                  log=lambda m, lv="info": 기록2.append((str(m), lv)))
    배치수 = -(-22 // 8)                       # 22줄을 8줄씩 → 3번
    봄(_가짜.부른수 == 배치수 + 1,
      "빈 답을 받고 그 배치만 한 번 더 물었다 — 나머지 배치는 그대로",
      f"모형 호출 {_가짜.부른수}번 (배치 {배치수} + 다시 1)")
    봄(any("쓸모없었습니다" in m for m, _ in 기록2), "왜 다시 물었는지 화면에 남긴다",
      next((m.strip()[:74] for m, _ in 기록2 if "쓸모없었습니다" in m), ""))
    bk2 = B.Book(res2["검사결과"], 시트)
    계2 = next(h for h in bk2.검증열 if "계획서" in h)
    값2 = str(bk2.ws[f"{bk2.검증열[계2]}3"].value or "")
    bk2.close()
    봄(값2 == "O", "다시 물어 받은 제대로 된 답이 칸에 들어갔다", f"3줄 계획서 = {값2}")
    _가짜.으쓱 = 0

    print()
    print("═" * 74)
    if 탈:
        print(f"❌ 어긋난 곳 {len(탈)}건")
        for x in 탈:
            print(f"   · {x}")
    else:
        print("✅ 폴더 하나와 시트 하나로 검사 칸까지 채워진다")
    print("═" * 74)
    print(f"   {res['검사결과']}")
    print(f"   {res['추출내역']}")
    return 1 if 탈 else 0


if __name__ == "__main__":
    raise SystemExit(main())
