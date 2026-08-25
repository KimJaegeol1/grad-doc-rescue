# -*- coding: utf-8 -*-
"""
smoke_paper — ⑦논문에서도 같은 엔진이 도는가 (성과유형 범용성)

여태 고친 것이 전부 **⑤인턴십** 하나를 돌려 보다 나왔다.  그러면 "인턴십용
특수처리를 다른 시트에 억지로 얹은 것 아니냐" 는 물음이 남는다.  ⑦논문은
성격이 아주 다르다 — 기업·기간·인원이 아니라 학술지·게재상태·저자·DOI 다.

여기서 깨지기 쉬운 것을 일부러 다 넣었다.

    저자 순서·영문 이름     Kwame Osei Mensah · 나다라 · J. Park
    학술지명 약칭          「환경분석학회지」 ↔ 「Korean J. Environ. Anal.」
    DOI                  10.1234/kjea.2025.31.2.87
    게재 상태             Published / Accepted / Online first
    게재일 정밀도          연월일 ↔ **연월까지만**  ← ⑤인턴십과 규칙이 같은지
    한 PDF 안의 여러 날짜   투고일 · 수정일 · 게재일 · 발행일
    제목의 특수문자        PM₂.₅ · 대소문자 · 하이픈
    사사 문구             본문 맨 뒤에만 있다

    python tools/smoke_paper.py
"""
from __future__ import annotations

import json
import shutil
import sys
import zipfile
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE.parent))

from openpyxl import Workbook, load_workbook                       # noqa: E402
from openpyxl.styles import PatternFill                            # noqa: E402

from core import ask as A, book as B, calc as C, flow as F, spec as S  # noqa: E402

일터 = HERE.parent / "workspace" / "_paper"
시트 = "⑦논문"
검증칠 = PatternFill("solid", fgColor="FCE4D6")
탈 = []


def 봄(ok, 글, 덧=""):
    print(f"  {'✓' if ok else '❌'} {글}" + (f"   {덧}" if 덧 else ""))
    if not ok:
        탈.append(글)


def 칸(제목):
    print("\n" + "═" * 74)
    print(제목)
    print("═" * 74)


# ══════════════════════════════════════════════════════════════
줄자료 = [
    # 논문명, 학술지명(약칭으로 적음), 저자, 게재일시
    ("PM₂.₅ 고농도 사례의 기여도 분석", "Korean J. Environ. Anal.",
     "나다라, Kwame Osei Mensah, 김한울", "2025-03-15"),
    ("도시 대기질 예측을 위한 machine-learning 기법 비교", "환경분석학회지",
     "라마바, J. Park", "2025-05"),                      # ★ 연월까지만
    ("생활폐기물 소각 시설의 미세먼지 배출 특성", "한국환경보건학회지",
     "사아자, 최푸름", "2025-07-01"),
    ("수용모델을 이용한 오염원 기여도 추정", "Atmospheric Environment",
     "가나다", "2026-02-10"),                            # ★ Accepted — 아직 게재 전
]


def 지어낸엑셀(경로: Path, 검증열들: list) -> int:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("논문 게재실적")                  # ★ 번호 기호가 없다
    ws["A1"] = "○○대학교 — 학술지 게재 실적"
    자료 = ["연번", "논문 제목", "게재지", "저자", "게재일자", "DOI"]
    #        └ 「논문 제목」·「게재지」·「저자」·「게재일자」 는 낱말장의 별칭
    for i, h in enumerate(자료, 1):
        ws.cell(3, i, h)                                # ★ 머리글이 3행
    for j, h in enumerate(검증열들, len(자료) + 1):
        ws.cell(3, j, h).fill = 검증칠
    ws.cell(3, len(자료) + len(검증열들) + 1, "비고").fill = 검증칠

    for r, (제목, 지, 저, 날) in enumerate(줄자료, 4):
        for c, v in enumerate((str(r - 3), 제목, 지, 저, 날,
                               f"10.1234/kjea.2025.31.{r}.87"), 1):
            ws.cell(r, c, v)
    wb.save(경로); wb.close()
    return len(줄자료)


def docx(path: Path, 줄들):
    body = "".join(f'<w:p><w:r><w:t xml:space="preserve">{x}</w:t></w:r></w:p>'
                   for x in 줄들)
    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "w") as z:
        z.writestr("[Content_Types].xml",
                   '<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats'
                   '.org/package/2006/content-types"><Default Extension="xml" '
                   'ContentType="application/xml"/><Override PartName="/word/'
                   'document.xml" ContentType="application/vnd.openxmlformats-'
                   'officedocument.wordprocessingml.document.main+xml"/></Types>')
        z.writestr("_rels/.rels",
                   '<?xml version="1.0"?><Relationships xmlns="http://schemas.'
                   'openxmlformats.org/package/2006/relationships"><Relationship '
                   'Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument'
                   '/2006/relationships/officeDocument" Target="word/document.xml"/>'
                   '</Relationships>')
        z.writestr("word/document.xml",
                   '<?xml version="1.0"?><w:document xmlns:w="http://schemas.'
                   'openxmlformats.org/wordprocessingml/2006/main"><w:body>'
                   + body + "</w:body></w:document>")


# ── 모형 흉내 ────────────────────────────────────────────────
class _응답:
    def __init__(self, d):
        self.usage = type("u", (), {"prompt_tokens": 999, "completion_tokens": 99})()
        self.choices = [type("c", (), {"message": type(
            "m", (), {"content": json.dumps(d, ensure_ascii=False)})()})()]


def _답(글: str, schema: dict) -> dict:
    줄꼴 = schema["properties"]["줄"]
    물은줄 = [int(k[1:]) for k in 줄꼴["properties"]]
    줄칸 = 줄꼴["properties"][f"r{물은줄[0]}"]["properties"]
    값이름 = list(줄칸["값"]["properties"])
    판정키 = [k for k in 줄칸 if k.startswith("v") and k[1:].isdigit()]
    파일 = [(ln[1:].partition("】")[0].strip(), ln[1:].partition("】")[2].strip())
            for ln in 글.splitlines() if ln.startswith("【file_")]

    #   행 → (파일 낱말, 게재상태, 게재일, 발행호 연월, 학술지 정식명, 저자)
    맺 = {4: ("논문1", "Published", "2025-03-15", "", "환경분석학회지",
              ["나다라", "Kwame Osei Mensah", "김한울"]),
          5: ("논문2", "Published", "", "2025-05", "환경분석학회지",
              ["라마바", "J. Park"]),                    # ★ 증빙도 연월까지만
          6: ("논문3", "Published", "2025-07-01", "", "한국환경보건학회지",
              ["사아자", "최푸름"]),
          7: ("논문4", "Accepted", "", "", "Atmospheric Environment",
              ["가나다"])}                               # ★ 아직 게재 전

    줄 = {}
    for n in 물은줄:
        낱말, 상태, 날, 연월, 지, 저자 = 맺.get(n, ("", "", "", "", "", []))
        f = [t for t, nm in 파일 if 낱말 and 낱말 in nm]
        값 = {k: "" for k in 값이름}
        for k, v in (("게재상태", 상태), ("날짜_Published", 날),
                     ("발행호_연월", 연월), ("학술지명", 지),
                     ("논문명", 줄자료[n - 4][0]),
                     ("사사_문구", "본 연구는 환경부 녹색융합기술인재양성사업의 지원을 받았음"),
                     (S.인원칸, str(len(저자)))):
            if k in 값:
                값[k] = v
        if S.명단칸 in 값:
            값[S.명단칸] = {"상태": "찾음" if 저자 else "명단없음", "이름": 저자}
        if "문서종류" in 값:
            값["문서종류"] = ["논문 원문", "게재 확인서"]
        한 = {"행": n, "맺은파일": f, "확신": "높음" if f else "못맺음",
              "맺은근거": "논문명·학술지 일치" if f else "폴더에서 못 찾음",
              "값": 값, "메모": ""}
        for k in 판정키:
            한[k] = {"검증열": k,
                     "판정": ("O" if f and 상태 == "Published" else "확인 불가"),
                     "출처": (f[0] if f else "없음"),
                     "근거": ("p.1 표제·서지사항에서 봄" if f else "맺은 서류가 없음"),
                     "비고": ""}
        줄[f"r{n}"] = 한
    쓴 = {y for x in 줄.values() for y in x["맺은파일"]}
    return {"줄": 줄, "안맺힌파일": [{"파일": t, "왜": "다른 시트 증빙인 듯하다"}
                                    for t, _n in 파일 if t not in 쓴]}


class _가짜:
    class chat:
        class completions:
            @staticmethod
            def create(**kw):
                return _응답(_답(kw["messages"][1]["content"],
                                kw["response_format"]["json_schema"]["schema"]))


# ══════════════════════════════════════════════════════════════
def main() -> int:
    import os
    if 일터.exists():
        shutil.rmtree(일터)
    일터.mkdir(parents=True)

    sp = S.load(F.기준찾기())
    검증열들 = [c["검증열"] for c in sp.항목(시트)]
    책 = 일터 / "○○대_성과관리현황.xlsx"
    줄수 = 지어낸엑셀(책, 검증열들)

    증빙 = 일터 / "증빙"
    for i, (제목, _지, 저, _날) in enumerate(줄자료, 1):
        docx(증빙 / f"논문{i}_원문.docx", [
            f"{제목}", f"저자: {저}",
            "Korean Journal of Environmental Analysis (환경분석학회지)"
            if i <= 2 else "학술지",
            "투고일: 2025-01-10", "수정일: 2025-02-20",      # ★ 여러 날짜
            f"게재확정일(Accepted): 2025-0{i}-01",
            f"게재일(Published): {_날}" if _날 else "게재 예정",
            f"DOI: 10.1234/kjea.2025.31.{i}.87",
            "사사: 본 연구는 환경부 녹색융합기술인재양성사업의 지원을 받았음"])
    docx(증빙 / "엉뚱한_인턴십_계획서.docx", ["인턴십 계획서", "기업명: 한빛환경기술"])

    A._client = lambda: _가짜()
    os.environ.setdefault("OPENAI_API_KEY", "시험용")

    칸("① ⑦논문을 끝까지 돌린다 — 인턴십과 필드가 완전히 다르다")
    print("  기업·기간·인원이 아니라 학술지·게재상태·저자·DOI 다.")
    print("  같은 엔진이 그대로 도는지가 이 시험의 전부다.\n")
    res = F.검사(책, 시트, 증빙, OCR=False, workers=2, log=lambda m, lv="info": None)
    봄(Path(res["검사결과"]).exists(), "검사 결과가 나왔다", Path(res["검사결과"]).name)

    칸("② 머리글 인식 — 세 숫자로 본다")
    bk = B.Book(res["검사결과"], 시트)
    지도, 못, 셈 = B.열지도(bk, sp.항목(시트))
    print(f"     그대로 {셈['그대로']} · 비슷 {셈['비슷']} · "
          f"모호 {셈['모호']} · 못 알아봄 {셈['못찾음']}\n")
    봄(셈["못찾음"] == 0 and 셈["모호"] == 0,
      "★ 「논문 제목·게재지·저자·게재일자」 를 낱말장으로 다 알아봤다",
      " · ".join(지도))
    줄들 = bk.줄들(지도)
    봄(len(줄들) == 줄수, f"논문 {줄수}줄을 다 읽는다", f"{len(줄들)}줄")
    봄(not bk.탈, "사라진 줄이 없다", str(bk.탈)[:60])
    bk.close()

    칸("③ 게재일 정밀도 — ⑤인턴십과 **같은 규칙**인가")
    print("  ⑦논문은 「연월까지만」 적힌 학술지가 흔하다.  ⑤인턴십에서 만든")
    print("  규칙(양쪽 다 온전하면 일자까지 · 한쪽만 연월이면 연월로)이")
    print("  시트가 달라져도 그대로 먹는지 본다.\n")
    항 = next(c for c in sp.항목(시트) if c["검증열"] == "게재일시 일치")

    def 날(엑, pub="", 연월=""):
        return C.재기(항, {"게재일시": 엑},
                      {"날짜_Published": pub, "발행호_연월": 연월},
                      맺은파일=["논문1.pdf"])

    사례 = [("2025-03-15", "2025-03-15", "", "O", "같은 날"),
            ("2025-03-15", "2025-03-20", "", "X", "연월 같아도 날이 다르면 X"),
            ("2025-05", "", "2025-05", "O", "양쪽 다 연월까지만"),
            ("2025-05-20", "", "2025-05", "O", "증빙이 연월까지만 — 봐준다"),
            ("2025-05", "2025-05-20", "", "O", "엑셀이 연월까지만 — 봐준다"),
            ("2025-05", "", "2025-09", "X", "연월이 다르면 X"),
            ("2026-02-10", "", "", "확인 불가", "아직 게재 전 — X 가 아니다")]
    for 엑, pub, 연월, 바람, 왜 in 사례:
        v, 비고, _ = 날(엑, pub, 연월)
        봄((v or "공란") == 바람, f"{바람:<6} {왜}",
          f"엑셀 {엑} ↔ 증빙 {pub or 연월 or '(없음)'}")

    칸("④ 저자 — 영문 이름과 이니셜이 섞여도 세어지는가")
    for 글, 바람 in [("나다라, Kwame Osei Mensah, 김한울", 3),
                     ("라마바, J. Park", 2),
                     ("사아자 · 최푸름", 2),
                     ("가나다", 1)]:
        봄(C.이름수(글) == 바람, f"「{글[:34]}」 → {바람}명", f"{C.이름수(글)}명")
    봄(C.명단수({S.명단칸: {"상태": "찾음", "이름": []}}) is None,
      "★ 「찾음」 인데 이름이 없으면 0 이 아니라 None — 거짓 X 를 막는다")

    칸("⑤ 검사 칸이 채워지고 자료 열은 그대로인가")
    wb = load_workbook(res["검사결과"])
    ws = wb["논문 게재실적"]
    쓴칸 = sum(1 for r in range(4, 4 + 줄수) for c in range(7, ws.max_column + 1)
               if ws.cell(r, c).value not in (None, ""))
    봄(쓴칸 > 0, f"검사 칸에 {쓴칸}개가 쓰였다")
    봄(str(ws.cell(4, 2).value or "").startswith("PM₂.₅"),
      "★ 제목의 특수문자를 한 글자도 안 건드렸다", str(ws.cell(4, 2).value or "")[:24])
    봄(str(ws.cell(4, 6).value or "").startswith("10.1234/"),
      "DOI 열도 그대로다", str(ws.cell(4, 6).value or ""))
    wb.close()

    print()
    print("═" * 74)
    if 탈:
        print(f"❌ 어긋난 곳 {len(탈)}건")
        for x in 탈:
            print(f"   · {x}")
    else:
        print("✅ ⑦논문에서도 같은 엔진이 그대로 돈다 — ⑤인턴십 전용이 아니다")
    print("═" * 74)
    return 1 if 탈 else 0


if __name__ == "__main__":
    raise SystemExit(main())
