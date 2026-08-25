# -*- coding: utf-8 -*-
r"""
stage3_write — 판정 보고 → 검증열 기입 (3단계 기입 층)
================================================================
경계: 기입만 한다.  판정하지 않는다.  판정 보고에 적힌 값을 그대로 옮긴다.

들어오는 것 (둘 다 파일 — 앞 단계의 메모리 상태에 기대지 않는다)
  · 3단계 판정 보고 JSON   시트 · 행 · 열 · 판정 · 비고
  · 성과관리현황 워크북     실제로 쓸 .xlsx
나가는 것
  · 검증열이 채워진 워크북 사본  (원본은 건드리지 않는다)

지키는 것
  · 검증열과 비고 열만 쓴다.  자료 열·학교 의견 열은 건드리지 않는다
  · 판정값은 넷뿐 — O · X · 확인 불가 · 공란(빈 문자열)
  · 비고는 [항목] 사유 꼴로 이어 붙인다.  이미 있던 비고는 지우지 않는다
  · 맑은 고딕 9pt · 가운데 정렬 · 색 채우기 없음   (0·1·2단계와 같은 표기)
  · 사람이 시트별로 미리 보고 승인한 뒤에 쓴다

설치:  pip install openpyxl

사용:
  python stage3_write.py --judge 판정보고.json --book 성과관리현황.xlsx
  python stage3_write.py --judge 판정보고.json --book 원본.xlsx -o 판정본.xlsx
  python stage3_write.py --judge 판정보고.json --book 원본.xlsx --dry-run
  python stage3_write.py --judge 판정보고.json --book 원본.xlsx --sheet 논문
"""

from __future__ import annotations

import argparse
import collections
import json
import os
import re
import shutil
import sys
import threading
import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import column_index_from_string as col_idx

if sys.platform == "win32":
    for _s in (sys.stdout, sys.stderr):
        try:
            _s.reconfigure(encoding="utf-8")
        except Exception:
            pass

FONT_NAME = "맑은 고딕"
FONT_SIZE = 9
O, X, UNK, BLANK = "O", "X", "확인 불가", ""

CANCEL = threading.Event()
_PREFIX = {"err": "❌ ", "warn": "⚠ ", "ok": "✓ ", "done": "★ ", "head": "■ ",
           "skip": "⏭ ", "info": ""}
_lock = threading.Lock()


def log(msg: str, level: str = "info"):
    print(_PREFIX.get(level, "") + str(msg), flush=True)


def log_many(lines):
    with _lock:
        for m, lv in lines:
            log(m, lv)


class Cancelled(Exception):
    pass


def _why(e: Exception) -> str:
    n, m = type(e).__name__, str(e)
    table = [
        ("PermissionError", "파일이 엑셀에서 열려 있습니다 — 닫고 다시 실행하세요"),
        ("FileNotFoundError", "파일을 찾을 수 없습니다"),
        ("KeyError", "워크북에 그 시트가 없습니다"),
        ("BadZipFile", "워크북이 손상되었습니다"),
        ("InvalidFileException", "엑셀 파일이 아니거나 형식이 맞지 않습니다"),
    ]
    for k, ko in table:
        if k in n:
            return ko
    return f"{n}: {m[:140]}"


# ══════════════════════════════════════════════════════════════
# 시트 찾기 — 이름이 조금 달라도 찾는다
# ══════════════════════════════════════════════════════════════
def _nk(s: str) -> str:
    return re.sub(r"[\s.·\-_()①②③④⑤⑥⑦⑧⑨⑩0-9]+", "", str(s or ""))


def find_sheet(wb, name: str):
    if name in wb.sheetnames:
        return wb[name]
    k = _nk(name)
    for t in wb.sheetnames:                       # 키워드로 (0·1·2단계와 같은 방식)
        if _nk(t) == k:
            return wb[t]
    for t in wb.sheetnames:
        if k and (k in _nk(t) or _nk(t) in k):
            return wb[t]
    return None


def verify_cols(ws, header_row: int) -> dict:
    """검증열(FCE4D6)과 비고 열을 찾는다 — 이름이 아니라 색으로(1단계 규약)."""
    VER = {"FFFCE4D6", "FCE4D6", "00FCE4D6"}
    out = {}
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(header_row, c)
        f = cell.fill
        rgb = getattr(getattr(f, "fgColor", None), "rgb", None)
        if f is not None and f.patternType == "solid" and isinstance(rgb, str) \
                and rgb.upper() in VER:
            out[cell.column_letter] = str(cell.value or "").replace("\n", " ").strip()
    return out


def base_remark(current: str, my_heads: set) -> str:
    """비고 칸의 '앞 단계가 남긴 부분' 만 골라낸다.

    3단계가 이미 쓴 대목([이 시트의 검증 항목] 로 시작하는 토막)은 걷어낸다.
    같은 워크북에 두 번 기입해도 3단계 사유가 겹겹이 쌓이지 않게 하기 위함이다
    (stage2_write 의 '스냅샷 기준으로 매번 새로 조립' 규약을 그대로 따른다).
    앞 단계(1·2단계)가 쓴 대목과 사람이 손으로 적은 글은 그대로 둔다.
    """
    cur = str(current or "").strip()
    if not cur:
        return ""
    keep = []
    # 줄바꿈으로도, 예전 꼴인 " / " 로도 가른다 — 지난번 비고가 남아 있을 수 있다
    for seg in re.split(r"\n+| / ", cur):
        t = seg.strip()
        if not t:
            continue
        m = re.match(r"^\[([^\]]+)\]", t)
        if m and m.group(1).strip() in my_heads:
            continue                      # 3단계가 지난번에 쓴 토막 — 버린다
        keep.append(t)
    return "\n".join(keep)


def remark_col(cols: dict) -> str:
    """검증열 중 '비고' 열.  '학교 의견' 은 절대 건드리지 않는다."""
    for c, h in cols.items():
        if "비고" in h and "학교" not in h:
            return c
    return ""


# ══════════════════════════════════════════════════════════════
# 기입 계획 — 무엇을 어디에 쓸지 먼저 만든다 (사람이 보고 승인)
# ══════════════════════════════════════════════════════════════
def plan(judge_path: Path, book_path: Path, sheet_filter: str = "") -> tuple[list, dict]:
    """반환: ([기입 항목], 집계).  워크북을 열어 보기만 하고 쓰지 않는다."""
    rep = json.loads(Path(judge_path).read_text("utf-8"))
    wb = load_workbook(book_path, data_only=False)
    stat = collections.Counter()
    items = []

    by_sheet = collections.defaultdict(list)
    for r in rep.get("행", []):
        by_sheet[r["시트"]].append(r)

    for sname, rows in by_sheet.items():
        if sheet_filter and _nk(sheet_filter) not in _nk(sname):
            continue
        ws = find_sheet(wb, sname)
        if ws is None:
            stat["시트없음"] += len(rows)
            log_many([(f"{sname}", "err"),
                      (f"     워크북에 이 시트가 없습니다 — 건너뜁니다", "info")])
            continue
        stat["시트"] += 1
        # 검증열 헤더행 — 색이 가장 많이 칠해진 줄
        hdr, best = 0, 0
        for rr in range(1, min(ws.max_row, 30) + 1):
            n = len(verify_cols(ws, rr))
            if n > best:
                hdr, best = rr, n
        cols = verify_cols(ws, hdr) if hdr else {}
        rc = remark_col(cols)

        heads = {c.get("헤더", "") for r in rows for c in r.get("열", []) if c.get("헤더")}
        for r in rows:
            row = r.get("행")
            if not row:
                stat["행번호없음"] += 1
                continue
            notes = []
            for c in r.get("열", []):
                letter, verdict, note = c.get("열"), c.get("판정"), c.get("비고", "")
                if not letter:
                    # 열이 없는 항목 = **비고에만 남길 말.**  칸은 건드리지 않는다.
                    #   예) 증빙에서 인원을 한 명도 못 뽑았다 — 어느 검증열에도
                    #       해당하지 않지만 사람이 알아야 하는 사실이다.
                    #   판정을 쓰지 않으므로 있던 값이 지워질 일도 없다.
                    if note:
                        notes.append(f"[{c.get('헤더', '')}] {note}")
                    continue
                if letter not in cols:                 # 색이 안 칠해진 열엔 쓰지 않는다
                    stat["검증열아님"] += 1
                    continue
                if "학교" in cols.get(letter, "") and "의견" in cols.get(letter, ""):
                    stat["학교의견_보호"] += 1
                    continue
                cur = ws.cell(row, col_idx(letter)).value
                items.append({
                    "시트": ws.title, "행": row, "열": letter,
                    "헤더": c.get("헤더", cols.get(letter, "")),
                    "판정": verdict if verdict is not None else BLANK,
                    "기존": "" if cur is None else str(cur),
                    "match_key": r.get("match_key", ""),
                })
                stat[verdict or "공란"] += 1
                if note:
                    notes.append(f"[{c.get('헤더','')}] {note}")
            if notes and rc:
                cur = ws.cell(row, col_idx(rc)).value
                items.append({
                    "시트": ws.title, "행": row, "열": rc, "헤더": "비고",
                    # ★ 줄바꿈으로 잇는다.  사유 안에 " / " 가 들어 있어서
                    #   ("엑셀 5명 / 이름 1명") 같은 기호로 이으면 어디서
                    #   토막이 끊기는지 안 보인다.  칸은 wrap_text 라 줄이 선다.
                    "판정": "\n".join(notes),
                    "기존": base_remark("" if cur is None else str(cur), heads),
                    "match_key": r.get("match_key", ""), "비고": True,
                })
                stat["비고"] += 1
    wb.close()
    return items, dict(stat)


# ══════════════════════════════════════════════════════════════
# 기입
# ══════════════════════════════════════════════════════════════
def run(judge_path: Path, book_path: Path, out_path: Path, *, sheet_filter="",
        overwrite=False, on_done=None) -> dict:
    """원본은 건드리지 않는다.  사본을 만들어 그 위에 쓴다."""
    items, stat = plan(judge_path, book_path, sheet_filter)
    out_path = Path(out_path)
    if out_path.resolve() != Path(book_path).resolve():
        shutil.copy2(book_path, out_path)

    wb = load_workbook(out_path, data_only=False)
    fnt = Font(name=FONT_NAME, size=FONT_SIZE)
    mid = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    wrote = skipped = 0
    by = collections.Counter()
    for i, it in enumerate(items, 1):
        if CANCEL.is_set():
            raise Cancelled()
        ws = wb[it["시트"]]
        cell = ws.cell(it["행"], col_idx(it["열"]))

        # 비고 — 앞 단계가 남긴 것 뒤에 이번 사유를 이어 붙인다.  덮어쓰지 않는다.
        #        계획을 세울 때 떠 둔 스냅샷으로 매번 새로 조립하므로,
        #        같은 워크북에 몇 번을 실행해도 결과가 같다.  길이는 자르지 않는다.
        if it.get("비고"):
            parts = [x for x in (it["기존"], it["판정"]) if x and str(x).strip()]
            cell.value = "\n".join(parts) if parts else None
            cell.font, cell.alignment = fnt, left
            wrote += 1
            by[it["시트"]] += 1
            if on_done:
                on_done(i, len(items), it)
            continue

        old = "" if cell.value is None else str(cell.value)
        if old.strip() and not overwrite:
            skipped += 1
            if on_done:
                on_done(i, len(items), it)
            continue
        cell.value = it["판정"] if it["판정"] else None
        cell.font = fnt
        cell.alignment = mid
        wrote += 1
        by[it["시트"]] += 1
        if on_done:
            on_done(i, len(items), it)

    wb.save(out_path)
    wb.close()
    stat["기입"] = wrote
    stat["건너뜀_기존값"] = skipped
    return {"산출": str(out_path), "집계": stat, "시트별": dict(by), "항목수": len(items)}


def summarize(items: list) -> dict:
    v = collections.Counter()
    for it in items:
        v["비고" if it.get("비고") else (it["판정"] or "공란")] += 1
    return dict(v)


def format_summary(res: dict, sec: float) -> list:
    st = res["집계"]
    out = [("─" * 66, "info"), (f"검증열 기입 완료 — {sec}초 걸렸습니다", "done")]
    out.append(("     " + "  ·  ".join(
        f"{k} {st[k]}" for k in (O, X, UNK, "공란", "비고") if st.get(k)), "info"))
    out.append((f"     기입 {st.get('기입',0)}칸"
                + (f"  ·  기존값이 있어 건너뜀 {st['건너뜀_기존값']}칸"
                   if st.get("건너뜀_기존값") else ""), "info"))
    if res.get("시트별"):
        out.append(("     시트별  " + "  ".join(f"{k} {v}" for k, v in res["시트별"].items()),
                    "info"))
    for k, ko in (("시트없음", "워크북에 없는 시트의 행"),
                  ("검증열아님", "검증열(주황 칠)이 아니어서 쓰지 않은 칸"),
                  ("학교의견_보호", "학교 의견 열이라 건드리지 않은 칸"),
                  ("행번호없음", "행 번호가 없는 판정")):
        if st.get(k):
            out.append((f"     {ko} {st[k]}", "warn"))
    out.append((f"     산출  {res['산출']}", "info"))
    return out


# ══════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════
def main() -> int:
    ap = argparse.ArgumentParser(description="3단계 기입 층 — 판정 보고를 검증열에 쓴다")
    ap.add_argument("--judge", required=True, help="3단계 판정 보고 JSON")
    ap.add_argument("--book", required=True, help="성과관리현황 워크북 .xlsx")
    ap.add_argument("-o", "--out", default="", help="산출 워크북 (기본: 원본 옆에 사본)")
    ap.add_argument("--sheet", default="", help="특정 시트만")
    ap.add_argument("--overwrite", action="store_true",
                    help="이미 값이 있는 칸도 덮어쓴다 (기본은 건너뜀)")
    ap.add_argument("--dry-run", action="store_true", help="쓰지 않고 계획만 보여준다")
    a = ap.parse_args()

    jp, bp = Path(a.judge).expanduser().resolve(), Path(a.book).expanduser().resolve()
    out = Path(a.out) if a.out else bp.with_name(bp.stem + "_3단계판정" + bp.suffix)

    items, stat = plan(jp, bp, a.sheet)
    print("═" * 70)
    print(f"  판정 보고  {jp.name}")
    print(f"  워크북     {bp.name}")
    print(f"  산출       {out.name}")
    print(f"  기입 대상  {len(items)}칸   {stat}")
    print("═" * 70)
    if not items:
        print("  쓸 것이 없습니다.")
        return 0
    if a.dry_run:
        cur = None
        for it in items[:60]:
            if it["시트"] != cur:
                cur = it["시트"]; print(f"\n  [{cur}]")
            mark = "  (기존값 있음)" if it["기존"].strip() else ""
            print(f"    {it['match_key']:<8} {it['열']}{it['행']:<4} "
                  f"{it['헤더'][:18]:<20} {str(it['판정'])[:40]}{mark}")
        if len(items) > 60:
            print(f"\n  … 외 {len(items)-60}칸")
        print("\n  실제로 쓰려면 --dry-run 을 빼고 실행하세요.")
        return 0

    t0 = time.time()
    res = run(jp, bp, out, sheet_filter=a.sheet, overwrite=a.overwrite)
    for m, lv in format_summary(res, round(time.time() - t0, 1)):
        print(_PREFIX.get(lv, "") + m)
    return 0


if __name__ == "__main__":
    if len(sys.argv) == 1:
        print("이 파일은 엔진입니다. 창으로 쓰려면:  python stage3_main.py")
        print("CLI:  python stage3_write.py --judge 판정보고.json --book 성과관리현황.xlsx")
        sys.exit(1)
    sys.exit(main())
