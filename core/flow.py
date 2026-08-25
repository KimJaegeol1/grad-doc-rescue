# -*- coding: utf-8 -*-
"""
core.flow — 여섯을 잇는다.  화면은 이 파일만 만난다.
================================================================
    ① 검수기준을 읽는다            core.spec
    ② 학교 엑셀을 연다             core.book
    ③ 증빙 폴더를 통째로 읽는다      engines.read  (폴더 안 폴더까지)
    ④ 시트를 통째로 모형에게 맡긴다   core.prompt · core.ask
    ⑤ 수와 날짜는 코드가 센다       core.calc
    ⑥ 검사 칸에 쓰고 내역을 뽑는다   engines.write · core.report

멈추는 자리
──────────
    검수기준이 없다        →  멈춘다.  규칙이 없으면 할 수 있는 일이 없다
    열쇠가 없다           →  멈춘다.  이 도구는 모형이 전부다
    검사 칸(주황)이 없다    →  멈춘다.  쓸 자리가 없다 — 검수도구로 먼저 만든다
    시트를 못 찾는다       →  멈춘다.  어느 시트가 있는지 알려 준다
    증빙이 한 장도 없다     →  멈춘다

나머지는 멈추지 않는다.  엑셀 열을 못 찾아도, 모형이 줄을 빠뜨려도, 글을
잘랐어도 **끝까지 가고 개요의 「봐야 할 것」 에 다 적는다.**  가다 말면
사람이 무엇이 됐고 무엇이 안 됐는지 알 수 없다.

★ 변하지 아니하는 다섯 (검수기준 개요 제0조) — 코드에서 어디가 지키나
─────────────────────────────────────────────────────────
    ① 없는 것은 만들지 않는다     prompt.표들·schema enum (파일은 표로만)
                                ask.검사 (없는 표·없는 줄번호를 버린다)
    ② 조용히 버리지 않는다        book.줄들 (값 있는 줄을 버리면 알린다)
                                book.열지도 (임자 없는 열을 알린다)
                                ask._한판 (빠진 줄·빈 답이면 다시 묻는다)
                                report._개요 「봐야 할 것」
    ③ 증빙이 이어져야 대조한다     ask._내리기 (맺은 파일 없는 O·X 를 내린다)
    ④ 같은 뜻끼리만 견준다        spec 낱말장·_겹침확인 (한 시트에서 두 항목이
                                같은 열을 노리면 막는다)
                                calc._수대조 (명단 수 대체는 뽑을값이 정할 때만)
    ⑤ 확실할 때만 O·X            calc 전반 (덜 뽑힘·정밀도 부족 → 확인 불가)
                                ask._내리기 (근거 없음·확신 낮음 → 확인 불가)

새 문제를 만나면 **새 규칙을 더하기 전에** 이 다섯 중 무엇을 제대로 못
지켰는지부터 본다.  여태 난 큰 사고는 전부 이 다섯 중 하나를 어긴 것이었다.

원본은 건드리지 않는다
────────────────────
검사 칸은 **사본**에 쓴다 (`engines.write` 규약).  그리고 쓰기 직전 원본을
따로 한 벌 더 떠 둔다 — 같은 파일을 두 번 돌릴 때를 위해서다.
"""

from __future__ import annotations

import json
import re
import shutil
import threading
from datetime import datetime
from pathlib import Path

from engines import read as R
from engines import write as W

from . import ask as A
from . import book as B
from . import calc as C
from . import report as RP
from . import spec as S


class 멈춤(Exception):
    """사람이 무엇을 해야 하는지 알려 주고 멈춘다."""


def _log(msg, level="info"):
    print(msg)


# ══════════════════════════════════════════════════════════════
def _중지잇기(cancel):
    """화면의 [중지] 를 엔진들에게 이어 준다.  반환: 걸음 사이에 부를 함수.

    엔진마다 제 CANCEL 이벤트를 갖고 있다 (engines.read · engines.write ·
    core.ask).  일부러 그렇게 두었다 — 엔진이 화면을 몰라야 하기 때문이다.
    그래서 잇는 일은 여기가 한다.

    ★ **시작할 때 반드시 clear() 한다.**  이벤트가 모듈 전역이라 지난번
      [중지] 가 남아 있으면 다음 실행이 시작하자마자 멈춘다.
    """
    엔진들 = [R.CANCEL, W.CANCEL, A.CANCEL]
    for e in 엔진들:
        e.clear()
    if cancel is None:
        return lambda: None

    def 지켜보기():
        cancel.wait()
        for e in 엔진들:
            e.set()

    threading.Thread(target=지켜보기, daemon=True).start()

    def 확인():
        if cancel.is_set():
            for e in 엔진들:
                e.set()
            raise A.Cancelled()

    return 확인


# ══════════════════════════════════════════════════════════════
본이름 = ("보조도구 검수기준.xlsx", "검수기준.xlsx")
곁장 = ("개요", "종합표", "추출 필드", "개정 이력")   # 검수기준에만 있는 장들


def _되살리기(이름: str) -> str:
    """깨진 이름을 되살려 본다.  못 되살리면 빈 글자.

    ★ 압축을 풀 때 한글이 깨지는 일이 잦다.  압축 파일이 UTF-8 표시 깃발을
      안 켜고 있으면 윈도우가 그 바이트를 CP949 로 풀어 버린다.

          보조도구 검수기준.xlsx   →   蹂댁“룄援 寃닔湲곗.xlsx

      파일은 멀쩡히 있는데 이름만 달라진 것이라, 사람은 폴더를 보며
      "있는데 왜 못 찾지" 한다.  그 자국을 되짚어 원래 이름을 보여 준다.
    """
    for 겉, 속 in (("cp949", "utf-8"), ("cp437", "utf-8"), ("latin-1", "utf-8")):
        try:
            t = 이름.encode(겉).decode(속)
        except Exception:                                # noqa: BLE001
            continue
        if t != 이름 and any("가" <= ch <= "힣" for ch in t):
            return t
    return ""


def _검수기준꼴(p: Path) -> int:
    """이 엑셀이 검수기준처럼 생겼나.  맞은 곁장 수 (0~4), 못 열면 -1.

    이름이 깨져도 **안은 안 깨진다.**  그래서 장 이름으로 알아본다.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(p, read_only=True, data_only=True)
        try:
            있는장 = {str(s).replace(" ", "") for s in wb.sheetnames}
        finally:
            wb.close()
    except Exception:                                    # noqa: BLE001
        return -1
    return sum(1 for g in 곁장 if g.replace(" ", "") in 있는장)


def _이건가요(자리: list) -> str:
    """찾아본 폴더의 엑셀을 훑어 「이건가요?」 를 만든다.

    그냥 "못 찾았습니다" 로 멈추면 사람은 다음에 뭘 해야 할지 모른다.
    폴더에 뭐가 있는지는 코드가 이미 볼 수 있으니, 보고 말해 주는 게 맞다.
    """
    후보 = []
    본 = set()
    for d in 자리:
        try:
            것들 = sorted(d.glob("*.xlsx"))
        except Exception:                                # noqa: BLE001
            continue
        for p in 것들:
            if p.name.startswith("~$") or str(p) in 본:
                continue
            본.add(str(p))
            후보.append((_검수기준꼴(p), p))
    if not 후보:
        return ""

    후보.sort(key=lambda x: -x[0])
    쪽 = ["", "혹시 이건가요?"]
    for 점수, p in 후보[:5]:
        원래 = _되살리기(p.name)
        쪽.append(f"  · {p.name}")
        if 점수 >= 2:
            쪽.append(f"      안을 열어 보니 검수기준이 맞습니다 (장 {점수}개 확인).")
        if 원래:
            쪽.append(f"      이름이 깨졌습니다.  원래 이름은 「{원래}」 입니다.")
            쪽.append("      압축을 풀 때 한글이 깨진 것입니다 — 이 이름으로 고쳐 주세요.")
        elif 점수 >= 2:
            쪽.append(f"      이름을 「{본이름[0]}」 로 고쳐 주세요.")
    if len(후보) > 5:
        쪽.append(f"  · … 그 밖 {len(후보) - 5}개")
    return "\n".join(쪽)


def 기준찾기(주어진="") -> Path:
    """검수기준 엑셀을 찾는다.  exe 옆 → docs/ 순.

    못 찾으면 그냥 멈추지 않는다.  찾아본 폴더의 엑셀을 훑어
    「혹시 이건가요?」 까지 말해 준다 (_이건가요).
    """
    자리 = []
    if 주어진:
        p = Path(주어진)
        if p.exists():
            return p
        자리.append(p.parent)
    else:
        try:
            import paths
            자리 += [Path(paths.exe_dir()), Path(paths.inner())]
        except Exception:                                # noqa: BLE001
            pass
        자리.append(Path(__file__).resolve().parent.parent / "docs")
        자리 += [d / "docs" for d in list(자리)]
    자리 = list(dict.fromkeys(d for d in 자리 if d and d.is_dir()))

    for d in 자리:
        for 이름 in 본이름:
            p = d / 이름
            if p.exists():
                return p

    쪽 = [f"검수기준 엑셀을 찾지 못했습니다 — {Path(주어진)}"] if 주어진 else [
        "검수기준 엑셀을 찾지 못했습니다.",
        f"「{본이름[0]}」 가 프로그램 옆에 있어야 합니다.",
        "이 파일이 검사 규칙 정본입니다 — 없으면 무엇을 볼지 알 수 없습니다."]
    쪽.append(_이건가요(자리))
    쪽 += ["", "찾아본 곳"] + [f"  · {d}" for d in 자리[:4]]
    raise 멈춤("\n".join(x for x in 쪽 if x is not None).rstrip())


def 시트고르기(book) -> list:
    """이 엑셀에서 검사할 수 있는 시트.  화면의 드롭다운을 여기서 채운다.

    검수기준이 아는 시트 가운데 **이 엑셀에 실제로 있고 검사 칸이 있는** 것만.
    고정 목록을 보여 주면 그 엑셀에 없는 것까지 고를 수 있고, 골라 봐야
    "시트가 없습니다" 로 끝난다.
    """
    sp = S.load(기준찾기())
    out = []
    for 이름 in sp.시트이름들():
        try:
            bk = B.Book(book, 이름)
        except B.못찾음:
            continue
        지도, _못, _셈 = B.열지도(bk, sp.항목(이름))
        줄 = bk.줄들(지도) if 지도 else []
        out.append({"본": 이름, "엑셀": bk.ws.title,
                    "줄수": len(줄), "검사칸": len(bk.검증열)})
        bk.close()
    return out


# ══════════════════════════════════════════════════════════════
def 검사(성과관리현황, 시트: str, 증빙폴더, *, 기준="", 캐시="", 낼곳="",
        모형="", workers=4, rps=2.0, 다시읽기=False, OCR=True,
        묶음=8, log=None, prog=None, cancel=None) -> dict:
    """끝까지 한 번에.  반환 {"검사결과", "추출내역", "집계", "신호"}.

    cancel  화면의 [중지] 이벤트.  `ui.runner` 가 언제나 넘긴다.
    묶음    한 번에 물을 줄 수.  줄이 서로 헷갈리는 시트면 줄여 본다.
    """
    log = log or _log
    멈춰 = _중지잇기(cancel)
    성과 = Path(성과관리현황)
    폴더 = Path(증빙폴더)
    캐시 = Path(캐시) if 캐시 else 폴더 / "text_cache"

    # ── ① 검수기준 ────────────────────────────────────────
    sp = S.load(기준찾기(기준))
    for 글, lv in sp.요약():
        log(글, lv)
    if 시트 not in sp.시트이름들():
        raise 멈춤(f"검수기준에 「{시트}」 시트가 없습니다.\n"
                   f"있는 것: {' · '.join(sp.시트이름들())}")

    멈춰()
    # ── ② 학교 엑셀 ───────────────────────────────────────
    log(f"[{시트}]  성과관리현황 엑셀을 엽니다", "head")
    try:
        bk = B.Book(성과, 시트)
    except B.못찾음 as e:
        raise 멈춤(str(e))
    if not bk.검증열:
        bk.close()
        raise 멈춤(
            f"「{시트}」 시트에 검사 칸(주황색)이 없습니다.\n"
            f"이 도구는 그 칸을 채우는 도구라 칸이 없으면 할 일이 없습니다.\n"
            f"검수도구로 [준비하기] 를 먼저 하신 뒤 그 파일을 넣어 주세요.")
    지도, 못찾은열, 열셈 = B.열지도(bk, sp.항목(시트))
    줄들 = bk.줄들(지도)
    명단 = bk.별첨1()
    엑셀값 = {r["행"]: r["값"] for r in 줄들}
    log(f"     워크북 「{bk.ws.title}」 · 머리글 {bk.머리글줄}행 · "
        f"검사 칸 {len(bk.검증열)}개 · 자료 줄 {len(줄들)}개 · 별첨1 {len(명단)}명", "ok")
    # ★ 「별칭이 몇 개냐」 가 아니라 **몇 개를 알아봤느냐** 가 범용성 지표다.
    임자알림 = []
    임자없는 = 열셈.get("임자없는열") or []
    말 = (f"머리글 인식 — 그대로 {열셈['그대로']} · 비슷 {열셈['비슷']} · "
          f"모호 {열셈['모호']} · 못 알아봄 {열셈['못찾음']}"
          + (f" · 임자 없는 열 {len(임자없는)}개" if 임자없는 else ""))
    log("     " + 말, "ok" if not (열셈["모호"] or 열셈["못찾음"] or 임자없는) else "warn")
    # ★ 개요 제0조 ② — 뜻을 모르는 열도 조용히 넘기지 않는다
    if 임자없는:
        책탈미리 = (f"어느 검수 항목도 보지 않는 열이 {len(임자없는)}개 있습니다 — "
                    f"「{'」 · 「'.join(임자없는[:6])}」"
                    f"{' …' if len(임자없는) > 6 else ''}.  "
                    f"검사 대상이 아니면 그대로 두시면 됩니다.  "
                    f"봐야 할 열인데 이름을 못 알아본 것이면 검수기준의 「열 낱말」 장에 "
                    f"한 줄 더해 주세요")
        log("     · " + 책탈미리, "skip")
        임자알림.append({"글": 책탈미리, "수준": "info"})
    for x in 못찾은열:
        if x.get("어떻게") == "모호":
            log(f"     ⚠ 「{x['검증열']}」 — 후보가 여럿이라 **안 골랐습니다**: "
                f"{x.get('후보') or '?'}.  검수기준의 「엑셀 대조 열」 에 "
                f"정확한 머리글을 적어 주세요", "warn")
        else:
            log(f"     ⚠ 「{x['검증열']}」 이 쓸 엑셀 열을 못 찾았습니다 — "
                f"찾아본 이름: {' · '.join(x['찾은이름'][:5])}"
                f"{' …' if len(x['찾은이름']) > 5 else ''}", "warn")
    # 학교가 적어 두었는데 우리가 아는 열엔 없어 건너뛴 줄 (core.book 이 담아 둔다)
    #   ★ 임자알림은 바로 위에서 이미 화면에 적었다.  여기서 또 적으면 같은
    #     말이 두 번 뜬다 (실제로 그랬다).  보고서로만 넘긴다.
    책탈 = [{"글": x, "수준": "warn"} for x in (bk.탈 or [])]
    for x in 책탈:
        log("     ⚠ " + x["글"], "warn")
    책탈 += 임자알림
    검증열자리 = _검증열자리(sp, 시트, bk.검증열, log)
    bk.close()
    if not 줄들:
        raise 멈춤(f"「{시트}」 시트에서 자료 줄을 하나도 읽지 못했습니다.\n"
                   f"머리글 아래에 값이 있는지 확인해 주세요.")

    멈춰()
    # ── ③ 증빙 읽기 ───────────────────────────────────────
    log("증빙 폴더를 통째로 읽습니다 — 폴더 안 폴더까지", "head")
    if not 폴더.exists():
        raise 멈춤(f"증빙 폴더가 없습니다 — {폴더}")
    할것, 건너뜀 = R.from_folder(폴더, recursive=True)
    if not 할것:
        raise 멈춤(f"증빙 폴더에서 읽을 수 있는 파일을 찾지 못했습니다 — {폴더}\n"
                   f"읽을 수 있는 것: {' · '.join(sorted(R.SUPPORTED))}")
    log(f"     읽을 것 {len(할것)}개 · 건너뜀 {len(건너뜀)}개", "ok")
    R.log = lambda m, lv="info": log("     " + str(m), lv)
    본것 = R.read_many(할것, 캐시, workers=workers, force=다시읽기,
                       allow_ocr=OCR, rps=rps,
                       on_done=(lambda i, n, r: prog and prog(i, n)))
    파일들 = _캐시읽기(캐시, log)
    if not 파일들:
        raise 멈춤("증빙에서 글자를 하나도 뽑지 못했습니다.\n"
                   "그림으로만 된 PDF 라면 Document AI 열쇠가 필요합니다 "
                   "(secrets/ 안 서비스계정 .json).")
    log(f"     글자를 뽑은 파일 {len(파일들)}개", "ok")

    멈춰()
    # ── ④ 모형 ────────────────────────────────────────────
    A.log = lambda m, lv="info": log(str(m), lv)
    A.열쇠읽기()
    ok, why = A.available()
    if not ok:
        raise 멈춤(why)
    res = A.run(sp, 시트, 줄들, 명단, 파일들, model=모형, prog=prog, 묶음=묶음)
    for 글, lv in A.format_summary(res):
        log(글, lv)

    멈춰()
    # ── ⑤ 코드가 셀 몫 ────────────────────────────────────
    본 = res["줄"]
    for 행, v in 본.items():
        v["엑셀값"] = 엑셀값.get(행, {})
        v["판정"] = C.채우기(sp, 시트, v, v["엑셀값"])

    멈춰()
    # ── ⑥ 쓰기 · 내역 ─────────────────────────────────────
    # ★ 검사 결과에도 **시트 이름을 붙인다.**
    # ────────────────────────────────────────
    #   engines.write 는 쓸 때마다 **원본을 다시 복사해** 그 위에 쓴다.
    #   그래서 이름이 하나뿐이면 ④세미나를 돌리는 순간 ⑤인턴십 결과가
    #   통째로 지워진다.  손으로 이름을 바꿔 두지 않으면 잃는다.
    #
    #   추출 내역은 처음부터 시트 이름이 붙어 있었는데 검사 결과만 없었다.
    #   둘을 맞춘다 — 시트마다 한 벌씩 남는다.
    #
    #       <학교>_보조_검사결과_⑤인턴십.xlsx
    #       <학교>_보조_추출내역_⑤인턴십.xlsx
    낼곳 = Path(낼곳) if 낼곳 else 성과.with_name(
        f"{_민(성과)}_보조_검사결과_{_안전(시트)}.xlsx")
    도장 = datetime.now().strftime("%y%m%d_%H%M%S")
    백업 = 성과.with_name(f"{성과.stem}_백업_{도장}.xlsx")
    shutil.copy2(성과, 백업)
    log(f"백업 — {백업.name}", "ok")

    보고 = _판정보고(시트, 본, 검증열자리)
    보고길 = 캐시.parent / "_보조_판정보고.json"
    보고길.write_text(json.dumps(보고, ensure_ascii=False, indent=1), "utf-8")
    W.log = lambda m, lv="info": log("     " + str(m), lv)
    쓴것 = W.run(보고길, 성과, 낼곳, sheet_filter=시트, overwrite=True)
    log(f"검사 칸을 채웠습니다 — {낼곳.name}", "done")

    내역길 = 성과.with_name(f"{_민(성과)}_보조_추출내역_{_안전(시트)}.xlsx")
    res["잰것"]["머리글인식"] = (f"그대로 {열셈['그대로']} · 비슷 {열셈['비슷']} · "
                                f"모호 {열셈['모호']} · 못 알아봄 {열셈['못찾음']}")
    RP.build(내역길, sp, 시트, 본, 신호=책탈 + res["신호"], 잰것=res["잰것"],
             자름=res["자름"], 안맺힌파일=res["안맺힌파일"],
             엑셀열이름=list(지도), 못찾은열=못찾은열)
    log(f"추출 내역을 뽑았습니다 — {내역길.name}", "done")

    센다: dict = {}
    for v in 본.values():
        for x in v["판정"].values():
            k = x["판정"] or "공란"
            센다[k] = 센다.get(k, 0) + 1
    return {"검사결과": str(낼곳), "추출내역": str(내역길), "백업": str(백업),
            "집계": {"엑셀 줄": len(줄들), "답한 줄": len(본),
                     "맺은 줄": sum(1 for v in 본.values() if v["맺은파일"]),
                     **센다},
            "신호": 책탈 + res["신호"]
                    + [{"글": f"「{x['검증열']}」 이 쓸 엑셀 열을 못 찾음",
                        "수준": "warn"} for x in 못찾은열]}


# ══════════════════════════════════════════════════════════════
def _검증열자리(sp, 시트, 워크북검증열: dict, log) -> dict:
    """검수기준의 검증열 이름 → 워크북의 열 문자.

    이름이 조금 어긋나도 찾는다.  못 찾으면 그 항목은 **쓰지 않는다** —
    엉뚱한 칸에 쓰느니 안 쓰는 게 낫다.
    """
    지도 = {B.nk(k): v for k, v in 워크북검증열.items()}
    out = {}
    for c in sp.항목(시트):
        이름 = c["검증열"]
        k = B.nk(이름)
        열 = 지도.get(k)
        if 열 is None:
            cand = [v for kk, v in 지도.items() if k and (k in kk or kk in k)]
            열 = cand[0] if len(cand) == 1 else None
        if 열 is None:
            log(f"     ⚠ 검사 칸 「{이름}」 을 워크북에서 못 찾았습니다 — "
                f"이 항목은 쓰지 않습니다", "warn")
            continue
        out[이름] = 열
    return out


_되뇜 = re.compile(r"^\s*\[([^\]]{1,40})\]\s*")


def _비고(이름: str, x: dict) -> str:
    """비고 칸에 남길 말.  ★ O 는 안 적는다.

    ────────────────────────────────────────────────────────────
    비고 칸은 검수자가 **무엇을 손봐야 하는지** 보는 자리다.  그런데 O 까지
    다 적어서 한 줄이 이렇게 됐다.

        [계획서 제출여부] [계획서 제출여부] 계획서 1건 / [결과보고서 요약서
        제출여부] [결과보고서 요약서 제출여부] 요약서 있음 / … / [참여인원
        일치] 엑셀 5명 / 이름 1명

    아홉 토막 중 정작 봐야 할 것은 맨 뒤 둘이다.  잘된 것을 적을수록 잘못된
    것이 묻힌다.  그래서 **X 와 확인 불가만** 적는다.

        X        학교에 반송이 나간다 — 왜 반송인지 적어야 한다
        확인 불가  사람이 봐야 한다 — 왜 못 가렸는지 적어야 한다
        O        할 일이 없다 — 안 적는다
        공란      판정할 대상이 아니다 — 안 적는다

    자세한 내력은 추출 내역 「판정」 장의 칸 메모에 그대로 남는다.  여기서
    줄이는 것은 학교 엑셀의 비고 칸뿐이다.

    되뇐 이름도 뗀다 — 모형이 「[계획서 제출여부] 계획서 1건」 처럼 항목
    이름을 앞에 한 번 적어 오는데, engines.write 가 또 붙여 두 번 찍힌다.
    """
    if (x.get("판정") or "") not in ("X", "확인 불가"):
        return ""
    t = str(x.get("비고") or "").strip()
    for _ in range(2):
        m = _되뇜.match(t)
        if not m or B.nk(m.group(1)) not in (B.nk(이름), B.nk(_짧은이름(이름))):
            break
        t = t[m.end():].strip()
    return t


def _짧은이름(검증열: str) -> str:
    """「계획서 제출여부」 → 「계획서」.  모형이 줄여 되뇌는 것도 잡는다."""
    t = re.sub(r"\s*\([^)]*\)\s*$", "", str(검증열 or "").strip())
    for 꼬리 in (" 제시 여부", " 포함 여부", " 제출여부", " 일치", " 여부", " 확인"):
        if t.endswith(꼬리):
            return t[: -len(꼬리)].strip() or 검증열
    return t or 검증열


def _판정보고(시트: str, 본: dict, 검증열자리: dict) -> dict:
    """engines.write 가 먹는 꼴로 옮긴다."""
    행들 = []
    for 행 in sorted(본):
        v = 본[행]
        칸 = []
        for 이름, x in (v["판정"] or {}).items():
            열 = 검증열자리.get(이름)
            if 열 is None:
                continue
            칸.append({"열": 열, "헤더": 이름, "판정": x["판정"] or "",
                       "비고": _비고(이름, x)})
        말 = _인원말(v)
        if 말:
            # 검증열이 **없는** 항목 = 비고에만 남길 말 (engines.write 규약)
            칸.append({"헤더": "인원", "비고": 말})
        if 칸:
            행들.append({"시트": 시트, "행": 행, "열": 칸,
                         "match_key": " · ".join(v["맺은파일"]) or "못 맺음"})
    return {"_만든이": "보조도구", "시트": 시트, "행": 행들}


def _인원말(v: dict) -> str:
    """증빙에서 사람을 한 명도 못 뽑았을 때 비고에만 남길 말.

    어느 검증열에도 해당하지 않지만 사람이 알아야 한다 — 대개 OCR 이 잘렸거나
    명단 쪽이 통째로 빠진 신호다.  판정을 안 쓰므로 있던 값이 지워질 일도 없다.
    """
    if not v.get("맺은파일"):
        return ""
    m = (v.get("값") or {}).get(S.명단칸)
    if isinstance(m, dict):
        상태 = m.get("상태")
        이름 = [x for x in (m.get("이름") or []) if str(x).strip()]
        if 상태 == "찾음" and 이름:
            return ""
        적힌 = str((v.get("값") or {}).get(S.인원칸, "") or "").strip()
        말 = {"명단없음": "증빙에 사람 이름이 실린 자리가 없음",
              "글자깨짐": "명단은 있으나 글자가 깨져 읽지 못함"}.get(
                  상태, "증빙에서 인원을 한 명도 뽑지 못함")
        return 말 + (f" — 문서에는 「{적힌[:20]}」 라고만 적혀 있음" if 적힌 else "")
    return ""


def _캐시읽기(캐시: Path, log) -> list:
    """engines.read 가 만든 텍스트 캐시를 [(원본 이름, 글)] 로.

    캐시 이름은 내용 주소라 `a1b2c3_보고서.txt` 꼴이다.  사람에게 보여 줄
    원래 이름은 옆의 `.meta.json` 에 `name` 으로 들어 있다.
    """
    out = []
    for txt in sorted(Path(캐시).glob("*.txt")):
        메타 = Path(str(txt)[:-4] + ".meta.json")     # 이름에 점이 있어도 안전하게
        이름 = txt.stem
        if 메타.exists():
            try:
                d = json.loads(메타.read_text("utf-8"))
                이름 = d.get("name") or Path(str(d.get("source", ""))).name or 이름
            except Exception:                        # noqa: BLE001
                pass
        try:
            글 = txt.read_text("utf-8", errors="replace")
        except Exception as e:                       # noqa: BLE001
            log(f"     {txt.name} 을 읽지 못했습니다 — {e}", "warn")
            continue
        if 글.strip():
            out.append((이름, 글))
    return out


def _민(p: Path) -> str:
    """앞 단계가 붙인 꼬리표를 뗀 학교 이름."""
    s = p.stem
    for 꼬리 in ("_검증본", "_1차_검사칸", "_2차_추출내역", "_3차_검사결과",
                 "_1단계판정", "_2단계판정", "_3단계판정",
                 "_보조_검사결과", "_보조_추출내역"):
        s = s.replace(꼬리, "")
    return s.strip() or p.stem


def _안전(s: str) -> str:
    for ch in '\\/:*?"<>|':
        s = s.replace(ch, "")
    return s
