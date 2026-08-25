# -*- coding: utf-8 -*-
"""
core.report — 추출 내역 엑셀.  **되짚는 길.**
================================================================
검사 칸에는 O·X·확인 불가 넉 자만 들어간다.  왜 그렇게 됐는지는 안 들어간다.
시트를 통째로 모형에게 맡기는 설계에서는 그 '왜' 를 잃으면 아무도 결과를
못 믿는다.  이 엑셀이 그 '왜' 다.

    개요       한눈에 · **봐야 할 것**
    <시트>     줄마다 두 줄 — 위는 엑셀에 적힌 값, 아래는 증빙에서 뽑은 값
    판정       줄 × 검증열 격자.  칸에 마우스를 올리면 근거가 나온다
    남은 것    못 맺은 줄 · 어느 줄에도 안 쓰인 파일

★ 표를 깨뜨리지 않는 법
──────────────────────
긴 글을 좁은 칸에 넣고 줄바꿈을 켜면 한 칸이 스무 줄이 되어 표가 무너진다
(전에 실제로 그랬다).  그래서
  · 표 **밖의** 글은 A열부터 넓게 병합하고 줄바꿈을 끈다
  · 표 **안의** 긴 글(근거)은 칸에 안 쓰고 **메모**로 붙인다
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as gl

글꼴 = "맑은 고딕"
F_TITLE = Font(name=글꼴, size=13, bold=True, color="1F3864")
F_HEAD = Font(name=글꼴, size=10, bold=True, color="FFFFFF")
F_BODY = Font(name=글꼴, size=10)
F_KEY = Font(name=글꼴, size=10, bold=True)
F_MUTED = Font(name=글꼴, size=9, color="7F7F7F")
F_SEC = Font(name=글꼴, size=11, bold=True, color="1F3864")

F_O = Font(name=글꼴, size=10, bold=True, color="1B7F3B")
F_X = Font(name=글꼴, size=10, bold=True, color="C62828")
F_UNK = Font(name=글꼴, size=10, color="B06000")
F_BLANK = Font(name=글꼴, size=10, color="A6A6A6")
F_부재 = Font(name=글꼴, size=10, color="BF8F00")
F_경고 = Font(name=글꼴, size=10, bold=True, color="C00000")

FILL_HEAD = PatternFill("solid", fgColor="4472C4")
FILL_엑셀 = PatternFill("solid", fgColor="F2F2F2")
FILL_증빙 = PatternFill("solid", fgColor="FFFFFF")
FILL_못맺 = PatternFill("solid", fgColor="FCE4D6")

얇 = Side(style="thin", color="D9D9D9")
BOX = Border(left=얇, right=얇, top=얇, bottom=얇)
A_L = Alignment(horizontal="left", vertical="center", wrap_text=False)
A_C = Alignment(horizontal="center", vertical="center", wrap_text=True)

부재 = "(부재)"
판정글꼴 = {"O": F_O, "X": F_X, "확인 불가": F_UNK, "": F_BLANK}


def _c(ws, r, c, v, font=F_BODY, fill=None, align=A_L, 메모=""):
    cell = ws.cell(r, c, v)
    cell.font = font
    cell.alignment = align
    cell.border = BOX
    if fill:
        cell.fill = fill
    if 메모:
        cell.comment = Comment(str(메모)[:900], "보조도구")
    return cell


def _글(ws, r, v, font=F_BODY, 폭=14, fill=None):
    """표 밖의 글 — 넓게 병합하고 줄바꿈은 끈다."""
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=폭)
    cell = ws.cell(r, 1, v)
    cell.font = font
    cell.alignment = A_L
    if fill:
        for c in range(1, 폭 + 1):
            ws.cell(r, c).fill = fill
    return cell


def _보임(v) -> str:
    """뽑은 값을 사람이 볼 글로.  참여명단은 상태까지 보여 준다."""
    if isinstance(v, dict):
        상태 = v.get("상태") or ""
        이름 = [str(x) for x in (v.get("이름") or []) if str(x).strip()]
        if 상태 and 상태 != "찾음":
            return f"({상태})"
        return ", ".join(이름) or 부재
    if isinstance(v, list):
        return ", ".join(str(x) for x in v if str(x).strip()) or 부재
    return str(v or "").strip() or 부재


# ══════════════════════════════════════════════════════════════
def build(out_path, sp, 시트: str, 본: dict, *,
          신호=None, 잰것=None, 자름=None, 안맺힌파일=None,
          엑셀열이름=None, 못찾은열=None) -> Path:
    """추출 내역 엑셀을 만든다.

    본        {행: {"엑셀값", "맺은파일", "맺은근거", "확신", "값", "판정", "메모"}}
    엑셀열이름 엑셀 쪽 표에 세울 열 이름 (core.book 의 열지도 키 차례)
    """
    신호 = 신호 or []
    잰것 = 잰것 or {}
    자름 = 자름 or []
    안맺힌파일 = 안맺힌파일 or []
    엑셀열이름 = list(엑셀열이름 or [])
    못찾은열 = 못찾은열 or []

    검증열 = [c["검증열"] for c in sp.항목(시트)]
    뽑을값 = sp.뽑을값들(시트)

    wb = Workbook()
    wb.remove(wb.active)
    _개요(wb, sp, 시트, 본, 신호, 잰것, 자름, 못찾은열)
    _값장(wb, sp, 시트, 본, 엑셀열이름, 뽑을값)
    _판정장(wb, sp, 시트, 본, 검증열)
    _남은것(wb, 본, 안맺힌파일)

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


# ── 개요 ──────────────────────────────────────────────────────
def _개요(wb, sp, 시트, 본, 신호, 잰것, 자름, 못찾은열):
    ws = wb.create_sheet("개요")
    ws.column_dimensions["A"].width = 22
    for c in range(2, 15):
        ws.column_dimensions[gl(c)].width = 11

    센다: dict = {}
    for v in 본.values():
        for x in v["판정"].values():
            k = x["판정"] or "공란"
            센다[k] = 센다.get(k, 0) + 1
    맺 = sum(1 for v in 본.values() if v["맺은파일"])

    r = 1
    _글(ws, r, f"{시트}   —   추출 내역", F_TITLE); r += 2
    머리 = ["엑셀 줄", "맺은 줄", "못 맺은 줄", "증빙 파일", "O", "X", "확인 불가", "공란"]
    값 = [len(본), 맺, len(본) - 맺, 잰것.get("파일", 0),
          센다.get("O", 0), 센다.get("X", 0), 센다.get("확인 불가", 0),
          센다.get("공란", 0)]
    for i, h in enumerate(머리, 1):
        _c(ws, r, i, h, F_HEAD, FILL_HEAD, A_C)
    for i, v in enumerate(값, 1):
        _c(ws, r + 1, i, v, F_KEY, align=A_C)
    r += 3

    # ── 봐야 할 것 ────────────────────────────────────────
    _글(ws, r, "▌ 봐야 할 것", F_SEC); r += 1
    볼것 = [s for s in 신호 if s.get("수준") in ("err", "warn")]
    참고 = [s for s in 신호 if s.get("수준") not in ("err", "warn")]
    if not 신호 and not 못찾은열 and not 자름:
        _글(ws, r, "   없습니다.", F_BODY); r += 1
    for s in 볼것:
        _글(ws, r, "   ❗ " + s["글"], F_경고); r += 1
    for x in 못찾은열:
        _글(ws, r, f"   ❗ 「{x['검증열']}」 이 쓸 엑셀 열을 못 찾았습니다 — "
                   f"찾아본 이름: {' · '.join(x['찾은이름'])}.  "
                   f"검수기준의 「엑셀 쪽 열」 칸에 실제 머리글을 더해 주세요",
            F_경고); r += 1
    for x in 자름:
        _글(ws, r, f"   ❗ {x['파일']} 은 {x['전체']:,}자 중 {x['실은것']:,}자만 "
                   f"모형에게 실었습니다 — 뒤쪽에만 있는 것은 못 본 것입니다",
            F_경고); r += 1
    for s in 참고:
        _글(ws, r, "   · " + s["글"], F_BODY); r += 1
    r += 1

    # ── 이 보고서를 보는 법 ───────────────────────────────
    _글(ws, r, "▌ 이 보고서를 보는 법", F_SEC); r += 1
    for ln in [
        "· 「값」 장은 줄마다 두 줄이다 — 위가 엑셀에 적힌 값, 아래가 증빙에서 뽑은 값.",
        "  나란히 놓았으므로 어디가 어긋났는지 눈으로 바로 보인다.",
        "· 「판정」 장의 칸에 **마우스를 올리면 근거**가 나온다 — 어디서 봤는지.",
        "  근거가 이 도구의 전부다.  시트를 통째로 모형에게 맡기므로, 되짚을 길이 그것뿐이다.",
        "· 근거에 「코드가 셈」 이 붙은 것은 모형이 아니라 코드가 판정한 것이다 "
        "(수·날짜·ISBN).",
        "· 근거에 「코드가 내림」 이 붙은 것은 모형이 O·X 를 냈으나 근거가 없거나 "
        "맺음이 헐거워 확인 불가로 내린 것이다.",
        "· (부재) 는 그 값을 못 뽑았다는 뜻이다.  틀렸다는 뜻이 아니다.",
        "· 판정은 검사 칸(성과관리현황 엑셀)에 쓰인다.  여기 판정 장은 그 사본이다.",
    ]:
        _글(ws, r, ln, F_BODY); r += 1
    r += 1

    if 잰것:
        _글(ws, r, "▌ 이번 실행", F_SEC); r += 1
        # ★ 나눔·다시부름·멈춘까닭을 함께 남긴다.  「왜 2번 배치만 빈 답이었나」 를
        #   결과만 보고는 못 짚어 화면 로그를 다시 받아야 했다.  그 일을 없앤다.
        for k in ("머리글인식", "나눔", "다시부름", "멈춘까닭", "어림토큰",
                  "쓴토큰_들어감", "쓴토큰_나옴", "채울칸",
                  "코드가셀칸", "안묻는칸", "명단"):
            if k in 잰것:
                v = 잰것[k]
                _글(ws, r, f"   {k:<12} "
                    + (f"{v:,}" if isinstance(v, int) else str(v)), F_MUTED); r += 1


# ── 값 장 ─────────────────────────────────────────────────────
def _값장(wb, sp, 시트, 본, 엑셀열이름, 뽑을값):
    ws = wb.create_sheet(_안전(시트)[:31])
    열 = ["줄", "어디 값", "맺은 파일", "확신"] + 엑셀열이름 + 뽑을값 + ["메모"]
    너비 = [6, 12, 34, 8] + [20] * len(엑셀열이름) + [20] * len(뽑을값) + [34]
    for i, h in enumerate(열, 1):
        _c(ws, 1, i, h, F_HEAD, FILL_HEAD, A_C)
        ws.column_dimensions[gl(i)].width = 너비[i - 1]
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "E2"

    r = 2
    for 행 in sorted(본):
        v = 본[행]
        맺 = " · ".join(v["맺은파일"]) or "어느 파일도 못 맺음"
        fill = None if v["맺은파일"] else FILL_못맺
        for 쪽 in ("엑셀", "증빙"):
            _c(ws, r, 1, 행 if 쪽 == "엑셀" else "", F_KEY, fill, A_C)
            _c(ws, r, 2, 쪽, F_MUTED, fill or (FILL_엑셀 if 쪽 == "엑셀" else None), A_C)
            _c(ws, r, 3, 맺 if 쪽 == "엑셀" else "", F_BODY, fill,
               메모=v["맺은근거"] if 쪽 == "엑셀" else "")
            _c(ws, r, 4, v["확신"] if 쪽 == "엑셀" else "", F_BODY, fill, A_C)
            for j, 이름 in enumerate(엑셀열이름, 5):
                글 = (v.get("엑셀값") or {}).get(이름, "") if 쪽 == "엑셀" else ""
                _c(ws, r, j, str(글).strip() or (부재 if 쪽 == "엑셀" else ""),
                   F_부재 if (쪽 == "엑셀" and not str(글).strip()) else F_BODY,
                   fill or (FILL_엑셀 if 쪽 == "엑셀" else None))
            for j, 이름 in enumerate(뽑을값, 5 + len(엑셀열이름)):
                글 = _보임((v.get("값") or {}).get(이름)) if 쪽 == "증빙" else ""
                _c(ws, r, j, 글,
                   F_부재 if (쪽 == "증빙" and 글.startswith("(")) else F_BODY,
                   fill or (FILL_엑셀 if 쪽 == "엑셀" else None))
            _c(ws, r, len(열), v.get("메모", "") if 쪽 == "증빙" else "",
               F_MUTED, fill)
            r += 1

    if not 본:
        _글(ws, 2, "모형이 아무 줄도 답하지 않았습니다.", F_경고)


# ── 판정 장 ───────────────────────────────────────────────────
def _판정장(wb, sp, 시트, 본, 검증열):
    ws = wb.create_sheet("판정")
    주체 = {c["검증열"]: c["주체"] for c in sp.항목(시트)}
    열 = ["줄", "맺은 파일"] + 검증열
    for i, h in enumerate(열, 1):
        _c(ws, 1, i, h, F_HEAD, FILL_HEAD, A_C)
        ws.column_dimensions[gl(i)].width = 6 if i == 1 else (
            30 if i == 2 else max(9, min(16, len(h) + 2)))
    ws.row_dimensions[1].height = 46
    ws.freeze_panes = "C2"

    r = 2
    for 행 in sorted(본):
        v = 본[행]
        _c(ws, r, 1, 행, F_KEY, align=A_C)
        _c(ws, r, 2, " · ".join(v["맺은파일"]) or "(못 맺음)",
           F_BODY, None if v["맺은파일"] else FILL_못맺)
        for j, h in enumerate(검증열, 3):
            x = (v["판정"] or {}).get(h)
            if x is None:
                _c(ws, r, j, "—", F_BLANK, align=A_C,
                   메모=f"{주체.get(h, '')} 주체인데 답이 없습니다")
                continue
            글 = x["판정"] or "공란"
            메모 = " / ".join(t for t in (x.get("비고"), x.get("근거")) if t)
            _c(ws, r, j, 글, 판정글꼴.get(x["판정"], F_BODY), align=A_C, 메모=메모)
        r += 1

    r += 1
    _글(ws, r, "칸에 마우스를 올리면 비고와 근거가 나옵니다.  "
               "「코드가 셈」 = 수·날짜를 코드가 판정한 것,  "
               "「코드가 내림」 = 모형이 낸 O·X 를 근거가 없어 확인 불가로 내린 것.",
        F_MUTED, 폭=max(3, len(열)))


# ── 남은 것 ───────────────────────────────────────────────────
def _남은것(wb, 본, 안맺힌파일):
    ws = wb.create_sheet("남은 것")
    r = 1
    _글(ws, r, "▌ 어느 파일도 못 맺은 줄", F_SEC, 폭=6); r += 1
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 90
    못 = [행 for 행 in sorted(본) if not 본[행]["맺은파일"]]
    if 못:
        _c(ws, r, 1, "줄", F_HEAD, FILL_HEAD, A_C)
        _c(ws, r, 2, "모형이 왜 못 맺었다 하는가", F_HEAD, FILL_HEAD, A_C)
        r += 1
        for 행 in 못:
            _c(ws, r, 1, 행, F_KEY, FILL_못맺, A_C)
            _c(ws, r, 2, 본[행]["맺은근거"] or 본[행].get("메모") or "(말이 없음)",
               F_BODY, FILL_못맺)
            r += 1
    else:
        _글(ws, r, "   없습니다 — 모든 줄이 파일을 찾았습니다.", F_BODY, 폭=6); r += 1
    r += 2

    _글(ws, r, "▌ 어느 줄에도 안 쓰인 파일", F_SEC, 폭=6); r += 1
    _글(ws, r, "   엑셀에 없는 실적이거나 다른 시트 증빙일 수 있습니다.  "
               "그것도 사람이 알아야 합니다.", F_MUTED, 폭=6); r += 1
    if 안맺힌파일:
        _c(ws, r, 1, "파일", F_HEAD, FILL_HEAD, A_C)
        _c(ws, r, 2, "모형이 하는 말", F_HEAD, FILL_HEAD, A_C)
        r += 1
        for x in 안맺힌파일:
            _c(ws, r, 1, str(x.get("파일", ""))[:60], F_BODY)
            _c(ws, r, 2, str(x.get("왜", "")), F_BODY)
            r += 1
    else:
        _글(ws, r, "   없습니다.", F_BODY, 폭=6)


def _안전(s: str) -> str:
    for ch in '\\/:*?"<>|[]':
        s = s.replace(ch, "")
    return s or "시트"
