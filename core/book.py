# -*- coding: utf-8 -*-
"""
core.book — 학교가 낸 성과관리현황 엑셀을 읽는다.
================================================================
읽어 오는 것 넷.

    ① 시트          이름이 어긋나도 찾는다 (⑧학술발표 ↔ 「⑦ 학술발표」)
    ② 검증열        우리가 채울 칸.  **색(FCE4D6)으로** 찾는다
    ③ 자료 열       학교가 적어 낸 값.  **검증열을 뺀 뒤** 찾는다   ★ 여기가 핵심
    ④ 별첨1 명단     참여학생 이름과 학생분류

★ 왜 '검증열을 뺀 뒤' 인가
─────────────────────────
검증열은 자료 열과 **같은 머리글 줄**에 끼워 넣은 것이다.  그리고 이름을
「자료열 + 일치」 로 짓는다.

    참여기관 일치   ⊃  참여기관          ⑤인턴십  시간 일치   ⊃  시간
    일시 일치      ⊃  일시             ⑤현장실습 교육인원 일치 ⊃  인원

그래서 이름만으로 찾으면 **자기가 채울 빈 칸을 자료 열로 집는다.**  값이
비었으니 "엑셀에 값이 없음" 이라 적고 판정을 공란으로 남긴다.  사람은 그것을
"판정을 안 했나 보다" 로 읽는다 — 열을 잘못 집었다는 것은 아무도 모른다.
실제로 서울대 건에서 이 때문에 다섯 칸이 조용히 비었다.

    ③프로젝트   참여기관 · 참여학생 수
    ⑤인턴십     시간
    ⑤현장실습   일시 · 인원

검증열이 어디인지는 이미 안다(색).  자료 열을 찾을 때 그 칸을 후보에서
빼기만 하면 된다.
"""

from __future__ import annotations

import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as gl

# ★ 괄호는 **지우지 않는다.**  「참여학생(명)」 은 사람 수, 「참여학생명」 은
#   이름이다.  괄호를 지우면 둘이 같아져 수 칸이 이름 칸을 집는다.
_PUNCT = re.compile(r"[\s·,\[\]{}~\-_.:;\"'’‘“”/\\「」]+")
_동그라미 = re.compile(r"^[①-⑳㉠-㉿]+")


def nk(s) -> str:
    """이름 견줄 때 쓰는 꼴.  공백·괄호·문장부호를 지운다."""
    return _PUNCT.sub("", unicodedata.normalize("NFC", str(s or ""))).lower()


def 민이름(s) -> str:
    """앞머리 번호 기호를 뗀 이름.  '⑧학술발표' → '학술발표'.

    학교가 번호를 잘못 매기는 일이 있다 — 서울대는 ⑧학술발표를
    「⑦ 학술발표」 로 냈다.  번호가 아니라 **이름으로** 찾아야 하는 이유다.
    """
    return nk(_동그라미.sub("", unicodedata.normalize("NFC", str(s or "")).strip()))


class 못찾음(Exception):
    pass


# ══════════════════════════════════════════════════════════════
class Book:
    """열어 놓은 성과관리현황 엑셀.

    ws        고른 시트
    머리글줄   자료 열과 검증열이 나란히 있는 줄
    검증열     {검증열 이름: 열문자}   — 색으로 찾은 것
    """

    def __init__(self, path, sheet: str):
        self.경로 = Path(path)
        self.찾는이름 = sheet
        self.wb = load_workbook(path, data_only=True)
        self.ws = _시트찾기(self.wb, sheet)
        if self.ws is None:
            있는 = " · ".join(self.wb.sheetnames)
            self.wb.close()
            raise 못찾음(
                f"엑셀에 「{sheet}」 시트가 없습니다.\n"
                f"들어 있는 시트: {있는}\n"
                f"이름이 조금 달라도 찾아보지만, 아주 다르면 못 찾습니다.")
        self.머리글줄 = _머리글줄(self.ws)
        self.검증열 = _검증열(self.ws, self.머리글줄)
        self.탈: list = []

    def close(self):
        try:
            self.wb.close()
        except Exception:                                # noqa: BLE001
            pass

    # ── ③ 자료 열 ─────────────────────────────────────────
    def 자료열(self, 이름들: list) -> tuple:
        """별칭을 앞에서부터 찾아본다.  반환 ([열번호…], 실제 머리글, 어떻게).

        어떻게   그대로 · 비슷 · 모호 · 못찾음      ← 실행 때 세어 보여 준다

        ★ 찾는 **방향**이 중요하다
        ─────────────────────────
            검수 항목 → 대표 이름 → 그 항목의 별칭 목록 → 엑셀 머리글 탐색

        반대로 하면 안 된다 — 엑셀 머리글 「일시」 를 보고 전역 낱말표에서
        무슨 뜻인지 되짚으면, 같은 「일시」 가 ④세미나에선 개최일이고
        ⑦논문에선 게재일이라 시트마다 딴 뜻이 된다.  **낱말표는 전역이어도
        뜻을 정하는 것은 언제나 그 검수 항목이다.**  그래서 이 함수는 낱말표를
        아예 모른다 — 이미 펴진 목록(이름들)만 받는다.  core.book 은 core.spec 을
        import 하지도 않는다.

        **검증열은 후보에서 뺀다.**  자기가 채울 칸을 자료 열로 집지 않기 위해서다.
        머리글 하나가 여러 칸에 걸쳐 있으면 **그 칸을 다 준다** (_자료머리글).
        """
        지도 = self._자료머리글()
        for 이름 in 이름들:
            k = nk(이름)
            if k in 지도:
                칸들 = 지도[k]
                return (칸들,
                        str(self.ws.cell(self.머리글줄, 칸들[0]).value or "").strip(),
                        "그대로")
        # 그래도 없으면 부분일치 — **후보가 하나일 때만** 쓴다
        모호 = []
        for 이름 in 이름들:
            k = nk(이름)
            cand = [(kk, v) for kk, v in 지도.items() if k and (k in kk or kk in k)]
            if len(cand) == 1:
                칸들 = cand[0][1]
                return (칸들,
                        str(self.ws.cell(self.머리글줄, 칸들[0]).value or "").strip(),
                        "비슷")
            if len(cand) > 1 and not 모호:
                # ★ 여럿이 걸리면 **안 고른다.**  아무거나 집으면 조용히 틀린다.
                #   다만 예전엔 그냥 넘겼다 — 이제 '모호' 로 세어 알린다.
                모호 = [str(self.ws.cell(self.머리글줄, v[0]).value or "").strip()
                        for _kk, v in cand][:4]
        return [], " · ".join(모호), ("모호" if 모호 else "못찾음")

    def _자료머리글(self) -> dict:
        """{정규화 머리글: [열번호…]}.  검증열은 빠져 있다.

        같은 이름이 여럿이면 **왼쪽 것**을 쓴다.  ⑥교재처럼 학교가 뒤에
        자체 확인란(교재명·저자·ISBN번호)을 또 만들어 두는 일이 있는데,
        앞쪽이 학교가 적어 낸 원자료다.

        ★ 값이 **여러 칸에 나뉜** 머리글이 있다
        ────────────────────────────────────
        서울대 ⑤인턴십은 「참여학생명」 이 O2:AE2 로 병합되어 있고, 이름이
        **한 칸에 하나씩** 들어 있다 (4줄은 18칸에 18명, 참여인원도 18).
        학교가 맞게 낸 것이다.

        맨 왼쪽 칸만 읽으면 이름이 1명이 되어 「참여인원 일치」 가

            코드가 셈 · 엑셀 18 ≠ 이름 1

        로 **전 줄 X** 가 된다.  X 는 학교에 반송이 나가는 판정이니, 멀쩡한
        실적 22건이 통째로 반송된다.  그래서 병합 범위를 끝까지 걷는다.

        칸을 넓히는 근거는 **실제 병합 범위뿐**이다.  머리글이 빈 칸을
        무턱대고 왼쪽 이름에 붙이면 엉뚱한 칸의 값이 섞여 든다.
        """
        if getattr(self, "_지도", None) is not None:
            return self._지도
        끝 = (self.ws.max_column or 0)
        검 = {c for c in range(1, 끝 + 1)
              if _검증칠(self.ws.cell(self.머리글줄, c))}
        폭 = {}                                   # 열번호 → (왼, 오른)  병합된 머리글
        for rng in self.ws.merged_cells.ranges:
            if rng.min_row <= self.머리글줄 <= rng.max_row:
                for cc in range(rng.min_col, rng.max_col + 1):
                    폭[cc] = (rng.min_col, rng.max_col)
        out = {}
        for c in range(1, 끝 + 1):
            if c in 검:
                continue
            v = self.ws.cell(self.머리글줄, c).value
            if v in (None, ""):
                continue
            왼, 오 = 폭.get(c, (c, c))
            if 오 > 왼:
                # ★ 병합 범위 밖으로 값이 새는 일이 있다.  서울대는 O2:AE2 로
                #   병합해 놓고 4줄의 18번째 이름을 AF 에 적었다.  이미 여러
                #   칸짜리라고 밝혀진 머리글이니, 머리글이 빈 칸이 이어지는
                #   데까지 따라간다 (다음 머리글·검증열에서 멈춘다).
                #   한 칸짜리 머리글은 이렇게 늘리지 않는다 — 엉뚱한 값이 섞인다.
                x = 오 + 1
                while (x <= 끝 and x not in 검
                       and self.ws.cell(self.머리글줄, x).value in (None, "")):
                    오, x = x, x + 1
            칸들 = [x for x in range(왼, min(오, 끝) + 1) if x not in 검]
            out.setdefault(nk(v), 칸들 or [c])    # 왼쪽 것이 이긴다
        self._지도 = out
        return out

    # ── 줄 읽기 ───────────────────────────────────────────
    def 줄들(self, 열지도: dict, 최대빈줄=20) -> list:
        """[{"행": 4, "값": {이름: 글}}].  병합된 칸은 이어받는다.

        성과관리현황은 한 건에 여러 줄이 붙고 기업명·일시 같은 값이 블록 첫
        줄에만 있다(나머지는 병합).  그냥 읽으면 둘째 줄부터 전부 빈 값이 되어
        "자료 행이 아니다" 하고 통째로 버려진다.

        ★ 열지도의 값은 **열 목록**이다.  「참여학생명」 처럼 이름이 한 칸에
          하나씩 든 머리글이 있어서다.  그런 칸은 쉼표로 이어 붙인다 —
          calc.이름수 가 쉼표·가운뎃점으로 가르므로 그대로 세어진다.
        """
        병합 = {}
        for rng in self.ws.merged_cells.ranges:
            머리 = self.ws.cell(rng.min_row, rng.min_col).value
            for rr in range(rng.min_row, rng.max_row + 1):
                for cc in range(rng.min_col, rng.max_col + 1):
                    병합[(rr, cc)] = 머리

        rows, 빈, 놓친 = [], 0, []
        for r in range(self.머리글줄 + 1, (self.ws.max_row or 0) + 1):
            값 = {}
            자기값 = False
            for 이름, 칸들 in 열지도.items():
                if 칸들 is None:
                    continue
                if isinstance(칸들, int):            # 옛 꼴도 받아 준다
                    칸들 = [칸들]
                조각 = []
                for c in 칸들:
                    v = self.ws.cell(r, c).value
                    if v not in (None, ""):
                        자기값 = True
                    elif len(칸들) == 1:
                        # 세로 병합을 이어받는 것은 **한 칸짜리 머리글에서만** 한다.
                        # 여러 칸짜리(참여학생명)에서 하면 블록 첫 줄의 이름이
                        # 아랫줄마다 되풀이돼 인원이 부풀려진다.
                        v = 병합.get((r, c))
                    if v not in (None, ""):
                        조각.append(str(v).strip())
                값[이름] = ", ".join(x for x in 조각 if x)
            if not any(값.values()):
                # ★ 우리가 아는 열엔 값이 없지만 **학교는 뭔가 적어 둔** 줄이 있다.
                #   그러면 그 줄은 조용히 사라진다 — 검사도 안 되고 알림도 없다.
                #   지어낸 학교로 돌려 보니 「참여기관」 을 못 찾는 바람에 5줄 중
                #   1줄이 그렇게 없어졌다.  버리되 **몇 줄을 버렸는지 남긴다.**
                if self._자료있나(r):
                    놓친.append(r)
                빈 += 1
                if 빈 >= 최대빈줄:
                    break
                continue
            if not 자기값 and not any(값.values()):
                continue
            빈 = 0
            rows.append({"행": r, "값": 값})
        if 놓친:
            self.탈.append(
                f"{len(놓친)}개 줄({', '.join(map(str, 놓친[:8]))}"
                f"{' …' if len(놓친) > 8 else ''})은 학교가 뭔가 적어 두었지만 "
                f"우리가 아는 열에는 값이 없어 건너뛰었습니다.  "
                f"검수기준의 「엑셀 대조 열」 칸에 그 머리글을 더해 주세요")
        return rows

    def _자료있나(self, r: int) -> bool:
        """이 줄에 학교가 적은 값이 있나 — 검사 칸(주황)은 빼고 본다."""
        검 = {gl(c) for c in range(1, (self.ws.max_column or 0) + 1)
              if _검증칠(self.ws.cell(self.머리글줄, c))}
        for c in range(1, (self.ws.max_column or 0) + 1):
            if gl(c) in 검:
                continue
            if self.ws.cell(r, c).value not in (None, ""):
                return True
        return False

    # ── ④ 별첨1 명단 ──────────────────────────────────────
    def 별첨1(self) -> list:
        """[{"이름": "김재걸", "학생분류": "재직자", ...}].  없으면 빈 목록.

        학생분류를 같이 읽는 이유 — ⑤인턴십 「재직자 증빙 제출여부(해당 시)」 가
        이것으로 해당 행인지 가린다.  모르면 확인 불가여야지 X 가 되면 안 된다.
        """
        ws = _시트찾기(self.wb, "별첨1 명단") or _시트찾기(self.wb, "별첨1")
        if ws is None:
            return []
        hr = _머리글줄(ws)
        머리 = {}
        for c in range(1, (ws.max_column or 0) + 1):
            v = str(ws.cell(hr, c).value or "").strip()
            if v:
                머리.setdefault(nk(v), (c, v))
        이름칸 = _고르기(머리, ["성명", "이름", "학생명", "참여학생"])
        분류칸 = _고르기(머리, ["학생분류", "구분", "재직", "유형"])
        if 이름칸 is None:
            return []
        out = []
        for r in range(hr + 1, (ws.max_row or 0) + 1):
            이름 = str(ws.cell(r, 이름칸).value or "").strip()
            if not (2 <= len(이름) <= 12) or 이름.isdigit():
                continue
            한 = {"이름": 이름}
            if 분류칸:
                한["학생분류"] = str(ws.cell(r, 분류칸).value or "").strip()
            out.append(한)
        # 같은 사람이 차년도마다 되풀이된다 (서울대는 112줄 중 실제 인원이 훨씬 적다).
        # 그대로 실으면 프롬프트만 길어지고 모형이 헷갈린다.  이름+분류로 한 번만 둔다.
        본, 봄 = [], set()
        for x in out:
            k = (x["이름"], x.get("학생분류", ""))
            if k in 봄:
                continue
            봄.add(k)
            본.append(x)
        return 본


# ══════════════════════════════════════════════════════════════
# 속
# ══════════════════════════════════════════════════════════════
_VER = {"FFFCE4D6", "FCE4D6", "00FCE4D6"}


def _검증칠(cell) -> bool:
    f = getattr(cell, "fill", None)
    rgb = getattr(getattr(f, "fgColor", None), "rgb", None)
    return bool(f is not None and f.patternType == "solid"
                and isinstance(rgb, str) and rgb.upper() in _VER)


def _시트찾기(wb, 이름: str):
    """이름이 어긋나도 찾는다.  ① 그대로 ② 번호를 뗀 이름 ③ 부분일치(하나뿐일 때)."""
    k = nk(이름)
    for t in wb.sheetnames:
        if nk(t) == k:
            return wb[t]
    m = 민이름(이름)
    맞 = [t for t in wb.sheetnames if 민이름(t) == m]
    if len(맞) == 1:
        return wb[맞[0]]
    맞 = [t for t in wb.sheetnames if m and (m in 민이름(t) or 민이름(t) in m)]
    return wb[맞[0]] if len(맞) == 1 else None


def _머리글줄(ws, scan=30) -> int:
    """검증열 색이 가장 많이 칠해진 줄.  칠이 없으면 값이 가장 많은 줄.

    글자 수만 세면 병합된 두 줄짜리 머리글이 자료 줄보다 칸이 적어, 자료 줄을
    머리글로 착각한다.  별첨3 개설교과는 머리글이 3행이고 나머지는 2행이라
    고정할 수도 없다.
    """
    색줄 = 색수 = 0
    글줄, 글수 = 1, 0
    for r in range(1, min(ws.max_row or 1, scan) + 1):
        칠 = 글 = 0
        for c in ws[r]:
            if _검증칠(c):
                칠 += 1
            if c.value not in (None, ""):
                글 += 1
        if 칠 > 색수:
            색줄, 색수 = r, 칠
        if 글 > 글수:
            글줄, 글수 = r, 글
    return 색줄 or 글줄


def _검증열(ws, hr: int) -> dict:
    """{머리글: 열문자}.  이름이 아니라 색으로 찾는다 (1단계 규약)."""
    out, 이어받기 = {}, ""
    for c in range(1, (ws.max_column or 0) + 1):
        cell = ws.cell(hr, c)
        v = cell.value
        if v not in (None, ""):
            이어받기 = str(v).strip()
        if _검증칠(cell) and 이어받기:
            out.setdefault(이어받기, gl(c))
    return out


def _고르기(머리: dict, 낱말들: list):
    for w in 낱말들:
        k = nk(w)
        if k in 머리:
            return 머리[k][0]
    for w in 낱말들:
        k = nk(w)
        cand = [v[0] for kk, v in 머리.items() if k and k in kk]
        if len(cand) == 1:
            return cand[0]
    return None


# ══════════════════════════════════════════════════════════════
def 열지도(bk: Book, spec_항목: list) -> tuple:
    """검수기준의 「엑셀 쪽 열」 을 실제 열 번호로 옮긴다.

    반환 ({보여줄이름: [열번호…]}, [못 찾은 것], 셈)
    보여줄이름은 검수기준에 적힌 **첫 별칭**(대표 이름)이다 — 프롬프트에서도 그 이름을 쓴다.
    값이 목록인 것은 「참여학생명」 처럼 한 머리글이 여러 칸에 걸치기 때문이다.

    셈  {"그대로", "비슷", "모호", "못찾음"}  ← 실행 때 이 세 숫자를 보여 준다.
        「별칭이 몇 개냐」 는 좋아졌다는 표시일 뿐 범용성의 증거가 아니다.
        새 학교에서 **몇 개를 알아봤고 몇 개를 못 알아봤는지**가 진짜 지표다.
    """
    지도, 못 = {}, []
    셈 = {"그대로": 0, "비슷": 0, "모호": 0, "못찾음": 0}
    for c in spec_항목:
        이름들 = c.get("엑셀열") or []
        if not 이름들:
            continue
        보임 = 이름들[0]
        if 보임 in 지도:
            continue
        칸들, 실제, 어떻게 = bk.자료열(이름들)
        셈[어떻게] = 셈.get(어떻게, 0) + 1
        if not 칸들:
            못.append({"검증열": c["검증열"], "찾은이름": 이름들,
                       "어떻게": 어떻게, "후보": 실제})
            continue
        지도[보임] = 칸들

    # ── ★ 어느 항목도 안 가져간 머리글 (개요 제0조 ②) ─────
    #   줄을 조용히 안 버리듯 **열도 조용히 안 넘긴다.**  학교가 「수행기관」
    #   이라 적었는데 우리가 뜻을 모르면, 넘기는 대신 "무엇과 이을지 확인
    #   필요" 라고 남긴다.  낱말장을 상상으로 늘리는 대신 이 알림을 보고
    #   **그 한 줄만** 더하는 것이 이 도구의 늘어나는 방식이다.
    쓴칸 = {c for v in 지도.values() for c in v}
    셈["임자없는열"] = []
    for k, 칸들 in bk._자료머리글().items():
        if set(칸들) & 쓴칸:
            continue
        머리 = str(bk.ws.cell(bk.머리글줄, 칸들[0]).value or "").strip()
        if 머리 and not _곁머리(머리):
            셈["임자없는열"].append(머리)
    return 지도, 못, 셈


_곁머리낱말 = ("번호", "연번", "순번", "no", "차년도", "분야", "학교", "비고",
              "학교 의견", "차수", "구분")


def _곁머리(머리: str) -> bool:
    """번호·차년도처럼 대조에 쓸 일이 없는 열은 알릴 것 없다."""
    k = nk(머리)
    return any(k == nk(w) or k.startswith(nk(w)) for w in _곁머리낱말)
